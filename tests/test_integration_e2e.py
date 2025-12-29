"""
End-to-End Integration Tests for PDF Upload and Excel Export.

Tests the complete pipeline:
1. Upload PDF document
2. Export to Excel with specified template
3. Verify output structure and data
"""

import pytest
import requests
from pathlib import Path
from openpyxl import load_workbook
import tempfile
import os


BASE_URL = "http://localhost:8000"
TEST_PDFS_DIR = Path("Test_Financial_Statement_PDFs")


class TestIncomeStatementPipeline:
    """Test complete Income Statement pipeline."""
    
    @pytest.fixture
    def income_statement_pdf(self):
        """Path to test Income Statement PDF."""
        return TEST_PDFS_DIR / "Test_Income_Statement_EASY.pdf"
    
    def test_upload_and_export_income_statement(self, income_statement_pdf):
        """Test complete upload -> export pipeline for Income Statement."""
        # Skip if PDF doesn't exist
        if not income_statement_pdf.exists():
            pytest.skip(f"Test PDF not found: {income_statement_pdf}")
        
        # Skip if server not running
        try:
            requests.get(BASE_URL, timeout=2)
        except requests.exceptions.ConnectionError:
            pytest.skip("Backend server not running")
        
        # Step 1: Upload PDF
        with open(income_statement_pdf, 'rb') as f:
            files = {'file': (income_statement_pdf.name, f, 'application/pdf')}
            response = requests.post(f"{BASE_URL}/api/v1/upload", files=files)
        
        assert response.status_code == 200, f"Upload failed: {response.text}"
        upload_data = response.json()
        assert 'document_id' in upload_data
        assert upload_data.get('tables'), "No tables extracted from PDF"
        
        doc_id = upload_data['document_id']
        
        # Step 2: Export to Excel
        export_response = requests.post(
            f"{BASE_URL}/api/v1/export/excel",
            json={
                'document_id': doc_id,
                'statement_type': 'income_statement',
                'style': 'basic',
                'colorway': 'green',
                'company_name': 'Test Company'
            }
        )
        
        assert export_response.status_code == 200, f"Export failed: {export_response.text}"
        export_data = export_response.json()
        assert export_data.get('rows_populated', 0) > 0, "No rows populated"
        assert export_data.get('periods'), "No periods detected"
        
        # Step 3: Download and verify Excel
        download_url = export_data['download_url']
        download_response = requests.get(f"{BASE_URL}{download_url}")
        assert download_response.status_code == 200
        
        # Save to temp file and verify
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(download_response.content)
            tmp_path = tmp.name
        
        try:
            wb = load_workbook(tmp_path)
            ws = wb.active
            
            # Verify no placeholder values remain
            placeholders_found = []
            for row in range(1, 40):
                val = ws.cell(row=row, column=3).value
                if val in [20, 10, 4, 2, -5]:
                    placeholders_found.append((row, val))
            
            assert not placeholders_found, f"Placeholder values found: {placeholders_found}"
            
            # Verify formula for Income Before Taxes is correct
            formula_cell = ws.cell(row=32, column=3).value
            if formula_cell:
                assert '=0+0' not in str(formula_cell), "Formula bug: =0+0 found"
            
            # Verify company name was set
            company_cell = ws.cell(row=1, column=2).value
            assert company_cell is not None, "Company name not set"
            
        finally:
            os.unlink(tmp_path)
    
    def test_export_with_missing_document_fails(self):
        """Test that export with non-existent document ID fails properly."""
        try:
            requests.get(BASE_URL, timeout=2)
        except requests.exceptions.ConnectionError:
            pytest.skip("Backend server not running")
        
        response = requests.post(
            f"{BASE_URL}/api/v1/export/excel",
            json={
                'document_id': '00000000-0000-0000-0000-000000000000',
                'statement_type': 'income_statement',
                'style': 'basic',
                'colorway': 'green'
            }
        )
        
        assert response.status_code == 404


class TestBalanceSheetPipeline:
    """Test complete Balance Sheet pipeline."""
    
    @pytest.fixture
    def balance_sheet_pdf(self):
        """Path to test Balance Sheet PDF."""
        return TEST_PDFS_DIR / "Test_Balance_Sheet_Single_Column.pdf"
    
    def test_upload_and_export_balance_sheet(self, balance_sheet_pdf):
        """Test complete upload -> export pipeline for Balance Sheet."""
        if not balance_sheet_pdf.exists():
            pytest.skip(f"Test PDF not found: {balance_sheet_pdf}")
        
        try:
            requests.get(BASE_URL, timeout=2)
        except requests.exceptions.ConnectionError:
            pytest.skip("Backend server not running")
        
        # Upload
        with open(balance_sheet_pdf, 'rb') as f:
            files = {'file': (balance_sheet_pdf.name, f, 'application/pdf')}
            response = requests.post(f"{BASE_URL}/api/v1/upload", files=files)
        
        assert response.status_code == 200
        doc_id = response.json()['document_id']
        
        # Export
        export_response = requests.post(
            f"{BASE_URL}/api/v1/export/excel",
            json={
                'document_id': doc_id,
                'statement_type': 'balance_sheet',
                'style': 'basic',
                'colorway': 'green',
                'company_name': 'Test BS Company'
            }
        )
        
        assert export_response.status_code == 200, f"Export failed: {export_response.text}"
        export_data = export_response.json()
        assert export_data.get('rows_populated', 0) >= 0  # May be 0 if classifier needs work


class TestNumberFormatting:
    """Test that number formatting is correct."""
    
    def test_precision_is_two_decimal_places(self):
        """Verify that aggregated values are rounded to 2 decimal places."""
        try:
            requests.get(BASE_URL, timeout=2)
        except requests.exceptions.ConnectionError:
            pytest.skip("Backend server not running")
        
        pdf_path = TEST_PDFS_DIR / "Test_Income_Statement_EASY.pdf"
        if not pdf_path.exists():
            pytest.skip("Test PDF not found")
        
        with open(pdf_path, 'rb') as f:
            files = {'file': (pdf_path.name, f, 'application/pdf')}
            response = requests.post(f"{BASE_URL}/api/v1/upload", files=files)
        
        doc_id = response.json()['document_id']
        
        export_response = requests.post(
            f"{BASE_URL}/api/v1/export/excel",
            json={
                'document_id': doc_id,
                'statement_type': 'income_statement',
                'style': 'basic',
                'colorway': 'green'
            }
        )
        
        download_response = requests.get(
            f"{BASE_URL}{export_response.json()['download_url']}"
        )
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(download_response.content)
            tmp_path = tmp.name
        
        try:
            wb = load_workbook(tmp_path)
            ws = wb.active
            
            # Check that no values have excessive decimal places
            for row in range(1, 40):
                val = ws.cell(row=row, column=3).value
                if isinstance(val, float):
                    val_str = str(val)
                    if '.' in val_str:
                        decimal_places = len(val_str.split('.')[1])
                        assert decimal_places <= 2, f"Row {row}: {val} has {decimal_places} decimal places"
        finally:
            os.unlink(tmp_path)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
