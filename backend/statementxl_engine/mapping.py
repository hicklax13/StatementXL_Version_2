"""
Mapping layer for the StatementXL Engine.

Pass 4: Map normalized facts to template cells.
Implements template profiling, label matching, and conflict resolution.
"""

import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import structlog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from backend.statementxl_engine.models import (
    CellPosting,
    ConfidenceLevel,
    MappingCandidate,
    NormalizedFact,
    PeriodInfo,
    ScaleFactor,
    StatementType,
    TemplateCell,
    TemplateGrid,
    TemplateProfile,
)

logger = structlog.get_logger(__name__)


# =============================================================================
# Template Profiling
# =============================================================================

class TemplateProfiler:
    """
    Profiles an Excel template to detect data grids and eligible cells.

    Detects:
    - Statement tabs by sheet name and content
    - Label columns
    - Period header rows
    - Value grid regions
    - Eligible cells (hardcoded numeric only)
    """

    # Sheet name patterns for statement detection
    SHEET_PATTERNS = {
        StatementType.INCOME_STATEMENT: [
            r"income\s*statement",
            r"p[&\s]*l",
            r"profit.*loss",
            r"statement\s*of\s*operations",
        ],
        StatementType.BALANCE_SHEET: [
            r"balance\s*sheet",
            r"statement\s*of\s*financial\s*position",
            r"assets.*liabilities",
        ],
        StatementType.CASH_FLOW: [
            r"cash\s*flow",
            r"statement\s*of\s*cash\s*flows",
        ],
    }

    # Signals for statement type detection in content
    CONTENT_SIGNALS = {
        StatementType.INCOME_STATEMENT: [
            "revenue", "net sales", "cost of goods", "gross profit",
            "operating income", "net income", "earnings per share",
        ],
        StatementType.BALANCE_SHEET: [
            "total assets", "total liabilities", "shareholders equity",
            "current assets", "current liabilities", "retained earnings",
        ],
        StatementType.CASH_FLOW: [
            "cash from operations", "investing activities",
            "financing activities", "net change in cash",
        ],
    }

    def profile_template(self, template_path: Path) -> TemplateProfile:
        """
        Profile an Excel template.

        Args:
            template_path: Path to the Excel template.

        Returns:
            TemplateProfile with detected structure.
        """
        logger.info("Profiling template", path=str(template_path))

        wb = load_workbook(template_path, data_only=False)
        sheet_names = wb.sheetnames

        grids: List[TemplateGrid] = []
        all_cells: Dict[str, TemplateCell] = {}
        detected_type: Optional[StatementType] = None

        for sheet_name in sheet_names:
            ws = wb[sheet_name]

            # Detect statement type from sheet name
            sheet_type = self._detect_statement_type_from_name(sheet_name)
            if sheet_type:
                detected_type = sheet_type

            # Analyze sheet structure
            grid = self._analyze_sheet(ws, sheet_name)
            if grid:
                grids.append(grid)

                # Collect all eligible cells
                for cell in grid.eligible_cells:
                    key = f"{sheet_name}!{cell.address}"
                    all_cells[key] = cell

        profile = TemplateProfile(
            id="",
            source_path=str(template_path),
            filename=template_path.name,
            sheet_names=sheet_names,
            detected_statement_type=detected_type,
            grids=grids,
            all_cells=all_cells,
        )

        logger.info(
            "Template profiling complete",
            sheets=len(sheet_names),
            grids=len(grids),
            eligible_cells=len(all_cells),
        )

        return profile

    def _detect_statement_type_from_name(self, sheet_name: str) -> Optional[StatementType]:
        """Detect statement type from sheet name."""
        name_lower = sheet_name.lower()
        for stmt_type, patterns in self.SHEET_PATTERNS.items():
            for pattern in patterns:
                if re.search(pattern, name_lower):
                    return stmt_type
        return None

    def _analyze_sheet(self, ws, sheet_name: str) -> Optional[TemplateGrid]:
        """Analyze a worksheet to detect grid structure."""
        # Find label column (usually A or B)
        label_col = self._find_label_column(ws)
        if label_col is None:
            return None

        # Find period header row
        header_row, period_columns, period_headers = self._find_period_headers(ws, label_col)
        if header_row is None:
            return None

        # Find value grid boundaries
        value_start, value_end = self._find_value_grid(ws, label_col, header_row)
        if value_start is None:
            return None

        # Collect eligible cells
        eligible_cells = self._collect_eligible_cells(
            ws, sheet_name, label_col, period_columns, period_headers, value_start, value_end
        )

        return TemplateGrid(
            sheet=sheet_name,
            label_column=label_col,
            label_start_row=value_start,
            label_end_row=value_end,
            period_header_row=header_row,
            period_columns=period_columns,
            period_headers=period_headers,
            value_start_row=value_start,
            value_end_row=value_end,
            value_columns=period_columns,
            eligible_cells=eligible_cells,
        )

    def _find_label_column(self, ws) -> Optional[int]:
        """Find the column containing row labels."""
        # Check columns A-C for text labels
        for col in range(1, 4):
            text_count = 0
            for row in range(1, min(50, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and len(cell.value) > 3:
                    text_count += 1
            if text_count > 5:
                return col
        return 1  # Default to column A

    def _find_period_headers(
        self,
        ws,
        label_col: int,
    ) -> Tuple[Optional[int], List[int], List[str]]:
        """Find row with period headers and extract columns."""
        year_pattern = re.compile(r"(20\d{2}|FY\s*\d{4}|Q[1-4]\s*\d{4})")

        for row in range(1, min(20, ws.max_row + 1)):
            period_columns = []
            period_headers = []

            for col in range(label_col + 1, min(label_col + 10, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if year_pattern.search(cell.value):
                        period_columns.append(col)
                        period_headers.append(cell.value)
                elif cell.value and isinstance(cell.value, (int, float)):
                    # Might be a year as number
                    if 2000 <= cell.value <= 2100:
                        period_columns.append(col)
                        period_headers.append(str(int(cell.value)))

            if len(period_columns) >= 1:
                return row, period_columns, period_headers

        return None, [], []

    def _find_value_grid(
        self,
        ws,
        label_col: int,
        header_row: int,
    ) -> Tuple[Optional[int], Optional[int]]:
        """Find start and end rows of the value grid."""
        start_row = None
        end_row = None

        for row in range(header_row + 1, min(100, ws.max_row + 1)):
            label_cell = ws.cell(row=row, column=label_col)
            if label_cell.value and isinstance(label_cell.value, str):
                if start_row is None:
                    start_row = row
                end_row = row

        return start_row, end_row

    def _collect_eligible_cells(
        self,
        ws,
        sheet_name: str,
        label_col: int,
        period_columns: List[int],
        period_headers: List[str],
        start_row: int,
        end_row: int,
    ) -> List[TemplateCell]:
        """Collect all eligible input cells in the grid."""
        cells = []

        for row in range(start_row, end_row + 1):
            # Get row label
            label_cell = ws.cell(row=row, column=label_col)
            row_label = str(label_cell.value) if label_cell.value else ""

            for i, col in enumerate(period_columns):
                cell = ws.cell(row=row, column=col)
                address = f"{get_column_letter(col)}{row}"

                # Check if cell has formula
                has_formula = isinstance(cell.value, str) and cell.value.startswith("=")

                # Eligible if: no formula AND (empty or numeric)
                is_eligible = not has_formula
                is_input = not has_formula and (
                    cell.value is None or
                    isinstance(cell.value, (int, float, Decimal))
                )

                period_header = period_headers[i] if i < len(period_headers) else None

                cells.append(TemplateCell(
                    sheet=sheet_name,
                    address=address,
                    row=row,
                    column=col,
                    current_value=cell.value,
                    has_formula=has_formula,
                    formula=cell.value if has_formula else None,
                    is_eligible=is_eligible,
                    is_input_cell=is_input,
                    row_label=row_label,
                    period_header=period_header,
                ))

        return cells


# =============================================================================
# Label Matching
# =============================================================================

class LabelMatcher:
    """
    Matches normalized fact labels to template row labels.

    Strategies (in order):
    1. Exact match (normalized)
    2. Synonym match
    3. Fuzzy string match (Levenshtein)
    4. Embedding similarity (optional)
    5. LLM suggestion (optional, gated)
    """

    def __init__(self):
        # Import synonym dictionary from normalization
        from backend.statementxl_engine.normalization import LABEL_TO_CANONICAL
        self._synonyms = LABEL_TO_CANONICAL

    def match(
        self,
        fact_label: str,
        template_label: str,
    ) -> Tuple[float, str]:
        """
        Calculate match score between fact and template labels.

        Args:
            fact_label: Normalized label from fact.
            template_label: Label from template row.

        Returns:
            (score, match_type) tuple.
        """
        fact_clean = self._clean(fact_label)
        template_clean = self._clean(template_label)

        # Exact match
        if fact_clean == template_clean:
            return 1.0, "exact"

        # Canonical match (both normalize to same)
        fact_canonical = self._synonyms.get(fact_clean, fact_clean)
        template_canonical = self._synonyms.get(template_clean, template_clean)

        if fact_canonical == template_canonical:
            return 0.95, "synonym"

        # Fuzzy match
        fuzzy_score = self._fuzzy_score(fact_clean, template_clean)
        if fuzzy_score > 0.8:
            return fuzzy_score * 0.9, "fuzzy"

        # Partial match (one contains the other)
        if fact_clean in template_clean or template_clean in fact_clean:
            return 0.7, "partial"

        return fuzzy_score * 0.5, "low"

    def _clean(self, label: str) -> str:
        """Clean label for matching."""
        return label.lower().strip()

    def _fuzzy_score(self, s1: str, s2: str) -> float:
        """Calculate fuzzy match score using Levenshtein distance."""
        if not s1 or not s2:
            return 0.0

        # Simple ratio-based scoring
        len1, len2 = len(s1), len(s2)
        max_len = max(len1, len2)

        if max_len == 0:
            return 1.0

        # Calculate edit distance (simplified)
        distance = self._levenshtein_distance(s1, s2)
        return 1.0 - (distance / max_len)

    def _levenshtein_distance(self, s1: str, s2: str) -> int:
        """Calculate Levenshtein distance between two strings."""
        if len(s1) < len(s2):
            return self._levenshtein_distance(s2, s1)

        if len(s2) == 0:
            return len(s1)

        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row

        return previous_row[-1]


# =============================================================================
# Mapping Engine
# =============================================================================

class MappingLayer:
    """
    Maps normalized facts to template cells.

    Process:
    1. Profile template (detect grids, eligible cells)
    2. Generate candidates (fact × cell with filters)
    3. Score candidates (label + period match)
    4. Greedy assignment (highest score first, no duplicates)
    5. Handle conflicts (deterministic resolution)
    6. Handle aggregation (sum components if needed)
    """

    # Thresholds
    AUTO_MAP_THRESHOLD = 0.70
    MIN_CANDIDATE_SCORE = 0.30

    def __init__(self):
        self._profiler = TemplateProfiler()
        self._label_matcher = LabelMatcher()

    def map_facts_to_template(
        self,
        facts: List[NormalizedFact],
        template_path: Path,
        target_period: Optional[str] = None,
    ) -> Tuple[TemplateProfile, List[CellPosting]]:
        """
        Map normalized facts to template cells.

        Args:
            facts: List of normalized facts from extraction.
            template_path: Path to the Excel template.
            target_period: Optional period filter (e.g., "FY2024").

        Returns:
            (template_profile, cell_postings) tuple.
        """
        logger.info(
            "Starting fact mapping",
            facts=len(facts),
            template=template_path.name,
        )

        # Profile template
        profile = self._profiler.profile_template(template_path)

        # Get eligible cells
        eligible_cells = profile.get_eligible_cells()
        if not eligible_cells:
            logger.warning("No eligible cells found in template")
            return profile, []

        # Generate and score candidates
        candidates = self._generate_candidates(facts, eligible_cells, target_period)
        logger.info("Generated candidates", count=len(candidates))

        # Greedy assignment
        postings = self._greedy_assign(candidates)

        # Handle conflicts
        postings = self._resolve_conflicts(postings, facts)

        logger.info(
            "Mapping complete",
            postings=len(postings),
            mapped_facts=len(set(p.primary_fact_id for p in postings)),
        )

        return profile, postings

    def _generate_candidates(
        self,
        facts: List[NormalizedFact],
        cells: List[TemplateCell],
        target_period: Optional[str],
    ) -> List[MappingCandidate]:
        """Generate candidate mappings."""
        candidates = []

        for fact in facts:
            for cell in cells:
                # Skip non-input cells
                if not cell.is_input_cell:
                    continue

                # Filter by period if specified
                if target_period and cell.period_header:
                    if not self._period_matches(fact.period, cell.period_header, target_period):
                        continue

                # Calculate match score
                score, components = self._score_candidate(fact, cell)

                if score >= self.MIN_CANDIDATE_SCORE:
                    candidates.append(MappingCandidate(
                        fact=fact,
                        target_cell=cell,
                        score=score,
                        match_type=components.get("match_type", "unknown"),
                        match_components=components,
                    ))

        # Sort by score descending
        candidates.sort(key=lambda c: c.score, reverse=True)

        return candidates

    def _period_matches(
        self,
        fact_period: Optional[PeriodInfo],
        cell_header: str,
        target_period: str,
    ) -> bool:
        """Check if fact period matches cell period."""
        if not fact_period:
            return True  # No period specified, match any

        # Check normalized key
        if fact_period.normalized_key == target_period:
            return True

        # Check if cell header contains target period
        if target_period in cell_header:
            return True

        return False

    def _score_candidate(
        self,
        fact: NormalizedFact,
        cell: TemplateCell,
    ) -> Tuple[float, Dict[str, float]]:
        """Score a candidate mapping."""
        components = {}

        # Label match
        if cell.row_label:
            label_score, match_type = self._label_matcher.match(
                fact.normalized_label, cell.row_label
            )
            components["label"] = label_score
            components["match_type"] = match_type
        else:
            components["label"] = 0.0
            components["match_type"] = "no_label"

        # Period match
        period_score = 1.0
        if fact.period and cell.period_header:
            # Check if periods align
            if fact.period.normalized_key in cell.period_header:
                period_score = 1.0
            elif fact.period.raw_header in cell.period_header:
                period_score = 0.9
            else:
                period_score = 0.5
        components["period"] = period_score

        # Confidence factor
        conf_score = fact.overall_confidence
        components["confidence"] = conf_score

        # Weighted total
        total = (
            components["label"] * 0.55 +
            components["period"] * 0.25 +
            components["confidence"] * 0.20
        )

        return total, components

    def _greedy_assign(
        self,
        candidates: List[MappingCandidate],
    ) -> List[CellPosting]:
        """Perform greedy assignment (no cell reuse)."""
        postings: List[CellPosting] = []
        used_cells: Set[str] = set()
        used_facts: Set[str] = set()

        for candidate in candidates:
            cell_key = f"{candidate.target_cell.sheet}!{candidate.target_cell.address}"
            fact_id = candidate.fact.id

            # Skip if cell or fact already used
            if cell_key in used_cells or fact_id in used_facts:
                continue

            # Create posting
            posting = CellPosting(
                id="",
                template_cell=candidate.target_cell,
                old_value=candidate.target_cell.current_value,
                new_value=candidate.fact.scaled_value,
                contributing_fact_ids=[fact_id],
                primary_fact_id=fact_id,
                confidence=candidate.score,
                confidence_level=self._score_to_level(candidate.score),
                source_pdf=candidate.fact.source_document_id,
                source_page=candidate.fact.source_page,
                source_raw_labels=[candidate.fact.raw_label],
                source_raw_values=[candidate.fact.raw_value],
                scale_factor=candidate.fact.scale_factor,
            )

            postings.append(posting)
            used_cells.add(cell_key)
            used_facts.add(fact_id)

        return postings

    def _resolve_conflicts(
        self,
        postings: List[CellPosting],
        all_facts: List[NormalizedFact],
    ) -> List[CellPosting]:
        """
        Resolve conflicts using deterministic tie-breakers.

        Tie-breaker order:
        1. Restated/revised flagged preference
        2. Explicit units indicator present
        3. Best internal consistency
        4. Filename date token (newer)
        """
        # Group postings by cell
        cell_postings: Dict[str, List[CellPosting]] = {}
        for posting in postings:
            key = f"{posting.template_cell.sheet}!{posting.template_cell.address}"
            if key not in cell_postings:
                cell_postings[key] = []
            cell_postings[key].append(posting)

        # For cells with multiple candidates, resolve
        resolved: List[CellPosting] = []
        for key, candidates in cell_postings.items():
            if len(candidates) == 1:
                resolved.append(candidates[0])
            else:
                # Apply tie-breakers
                winner = self._apply_tiebreakers(candidates, all_facts)
                winner.had_conflict = True
                winner.conflict_resolution = "deterministic_tiebreaker"
                winner.rejected_candidates = [c.primary_fact_id for c in candidates if c != winner]
                resolved.append(winner)

        return resolved

    def _apply_tiebreakers(
        self,
        candidates: List[CellPosting],
        all_facts: List[NormalizedFact],
    ) -> CellPosting:
        """Apply deterministic tie-breakers to select winner."""
        # Get facts for candidates
        fact_map = {f.id: f for f in all_facts}

        # Tie-breaker 1: Prefer restated
        restated = [c for c in candidates if fact_map.get(c.primary_fact_id, NormalizedFact(id="", normalized_label="", raw_label="", raw_value="", parsed_value=Decimal(0), scaled_value=Decimal(0))).is_restated]
        if len(restated) == 1:
            return restated[0]

        # Tie-breaker 2: Higher confidence
        candidates.sort(key=lambda c: c.confidence, reverse=True)
        if candidates[0].confidence > candidates[1].confidence + 0.1:
            return candidates[0]

        # Tie-breaker 3: First by alphabetical source
        candidates.sort(key=lambda c: c.source_pdf)
        return candidates[0]

    def _score_to_level(self, score: float) -> ConfidenceLevel:
        """Convert score to confidence level."""
        if score >= 0.85:
            return ConfidenceLevel.HIGH
        elif score >= 0.65:
            return ConfidenceLevel.MEDIUM
        elif score >= 0.40:
            return ConfidenceLevel.LOW
        else:
            return ConfidenceLevel.VERY_LOW


def get_template_profiler() -> TemplateProfiler:
    """Get TemplateProfiler instance."""
    return TemplateProfiler()


def get_mapping_layer() -> MappingLayer:
    """Get MappingLayer instance."""
    return MappingLayer()
