"""
Audit sheet generation for the StatementXL Engine.

Pass 7: Generate comprehensive audit sheet with full lineage.
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional

import structlog
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.statementxl_engine.models import (
    CellPosting,
    ConfidenceLevel,
    DocumentEvidence,
    NormalizedFact,
    ReconciliationResult,
    RunAudit,
    ScaleFactor,
    StatementSection,
    TemplateProfile,
)

logger = structlog.get_logger(__name__)


class AuditSheetGenerator:
    """
    Generates the mandatory Audit sheet in the output workbook.

    Contents:
    1. Metadata: System, timestamp, config, template, PDF list
    2. Statement sections per PDF with classification rationale
    3. Units/scale by section
    4. Period mapping table (raw → normalized → end_date → duration)
    5. Lineage table with full column set
    6. Exceptions and warnings
    7. Reconciliation results
    """

    # Column headers for lineage table
    LINEAGE_COLUMNS = [
        "TemplateTab",
        "TemplateLineItem",
        "CellAddress",
        "TemplatePeriodHeader",
        "NormalizedPeriodKey",
        "SourcePDF",
        "SourceSectionTitle",
        "Page#",
        "SourceRawLabel(s)",
        "RawValue(s)",
        "ScaleFactor",
        "AggregationComponents",
        "AggregationFormula",
        "FinalPostedValue($)",
        "Confidence(H/M/L)",
        "ConflictFlag(Y/N)",
        "Notes",
    ]

    # Styles
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    SECTION_FONT = Font(bold=True)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    def generate_audit_sheet(
        self,
        workbook_path: Path,
        audit: RunAudit,
    ) -> None:
        """
        Add Audit sheet to the output workbook.

        Args:
            workbook_path: Path to the output workbook.
            audit: Complete run audit data.
        """
        logger.info("Generating Audit sheet", workbook=workbook_path.name)

        wb = load_workbook(workbook_path)

        # Remove existing Audit sheet if present
        if "Audit" in wb.sheetnames:
            del wb["Audit"]

        # Create new Audit sheet
        ws = wb.create_sheet("Audit")

        current_row = 1

        # Section 1: Metadata
        current_row = self._write_metadata_section(ws, audit, current_row)

        # Section 2: Statement Sections
        current_row = self._write_sections_section(ws, audit, current_row)

        # Section 3: Units/Scale
        current_row = self._write_scale_section(ws, audit, current_row)

        # Section 4: Period Mapping
        current_row = self._write_period_section(ws, audit, current_row)

        # Section 5: Lineage Table
        current_row = self._write_lineage_section(ws, audit, current_row)

        # Section 6: Exceptions
        current_row = self._write_exceptions_section(ws, audit, current_row)

        # Section 7: Reconciliation
        current_row = self._write_reconciliation_section(ws, audit, current_row)

        # Auto-fit columns
        self._auto_fit_columns(ws)

        # Save
        wb.save(workbook_path)

        logger.info("Audit sheet generated", rows=current_row)

    def _write_section_header(
        self,
        ws,
        row: int,
        title: str,
    ) -> int:
        """Write a section header."""
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.font = self.SECTION_FONT
        cell.fill = self.SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        return row + 1

    def _write_metadata_section(
        self,
        ws,
        audit: RunAudit,
        start_row: int,
    ) -> int:
        """Write metadata section."""
        row = self._write_section_header(ws, start_row, "METADATA")

        metadata = [
            ("System", audit.system),
            ("Engine Version", audit.engine_version),
            ("Run ID", audit.run_id),
            ("Timestamp", audit.timestamp.isoformat()),
            ("Template", audit.template_filename),
            ("PDFs Processed", ", ".join(audit.pdf_filenames)),
            ("Statement Type", audit.statement_type.value if audit.statement_type else "Auto-detected"),
            ("Total Facts Extracted", audit.total_facts),
            ("Facts Mapped", audit.mapped_facts),
            ("Cells Posted", audit.posted_cells),
        ]

        for label, value in metadata:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1

        return row + 1

    def _write_sections_section(
        self,
        ws,
        audit: RunAudit,
        start_row: int,
    ) -> int:
        """Write statement sections detected."""
        row = self._write_section_header(ws, start_row, "DETECTED STATEMENT SECTIONS")

        # Headers
        headers = ["PDF", "Section Title", "Statement Type", "Page", "Rows", "Method", "Confidence", "Rationale"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        row += 1

        # Data
        for section in audit.statement_sections:
            ws.cell(row=row, column=1, value=section.source_table_id.split("_")[0] if "_" in section.source_table_id else "")
            ws.cell(row=row, column=2, value=section.title)
            ws.cell(row=row, column=3, value=section.statement_type.value)
            ws.cell(row=row, column=4, value=section.page)
            ws.cell(row=row, column=5, value=f"{section.start_row}-{section.end_row}")
            ws.cell(row=row, column=6, value=section.classification_method)
            ws.cell(row=row, column=7, value=f"{section.confidence:.2%}")
            ws.cell(row=row, column=8, value=section.rationale)
            row += 1

        return row + 1

    def _write_scale_section(
        self,
        ws,
        audit: RunAudit,
        start_row: int,
    ) -> int:
        """Write scale factor section."""
        row = self._write_section_header(ws, start_row, "UNITS / SCALE FACTORS")

        # Headers
        headers = ["Location", "Scale Factor", "Multiplier"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        row += 1

        # Data
        for location, scale in audit.detected_scale_factors.items():
            ws.cell(row=row, column=1, value=location)
            ws.cell(row=row, column=2, value=scale.name)
            ws.cell(row=row, column=3, value=scale.value)
            row += 1

        if not audit.detected_scale_factors:
            ws.cell(row=row, column=1, value="(All values in units - no scaling detected)")
            row += 1

        return row + 1

    def _write_period_section(
        self,
        ws,
        audit: RunAudit,
        start_row: int,
    ) -> int:
        """Write period mapping section."""
        row = self._write_section_header(ws, start_row, "PERIOD MAPPING")

        # Headers
        headers = ["Raw Header", "Normalized Key", "End Date", "Duration (Months)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        row += 1

        # Data
        for mapping in audit.period_mappings:
            ws.cell(row=row, column=1, value=mapping.get("raw_header", ""))
            ws.cell(row=row, column=2, value=mapping.get("normalized_key", ""))
            ws.cell(row=row, column=3, value=mapping.get("end_date", ""))
            ws.cell(row=row, column=4, value=mapping.get("duration_months", ""))
            row += 1

        if not audit.period_mappings:
            ws.cell(row=row, column=1, value="(No period mappings)")
            row += 1

        return row + 1

    def _write_lineage_section(
        self,
        ws,
        audit: RunAudit,
        start_row: int,
    ) -> int:
        """Write lineage table section."""
        row = self._write_section_header(ws, start_row, "LINEAGE TABLE")

        # Headers
        for col, header in enumerate(self.LINEAGE_COLUMNS, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        row += 1

        # Data
        for posting in audit.cell_postings:
            cell_info = posting.template_cell
            period_key = cell_info.normalized_period or ""

            # Confidence level
            conf_str = {
                ConfidenceLevel.HIGH: "H",
                ConfidenceLevel.MEDIUM: "M",
                ConfidenceLevel.LOW: "L",
                ConfidenceLevel.VERY_LOW: "VL",
            }.get(posting.confidence_level, "?")

            data = [
                cell_info.sheet,  # TemplateTab
                cell_info.row_label or "",  # TemplateLineItem
                cell_info.address,  # CellAddress
                cell_info.period_header or "",  # TemplatePeriodHeader
                period_key,  # NormalizedPeriodKey
                posting.source_pdf,  # SourcePDF
                posting.source_section_title,  # SourceSectionTitle
                posting.source_page,  # Page#
                ", ".join(posting.source_raw_labels),  # SourceRawLabel(s)
                ", ".join(posting.source_raw_values),  # RawValue(s)
                posting.scale_factor.name,  # ScaleFactor
                ", ".join(posting.aggregation_components),  # AggregationComponents
                posting.aggregation_formula or "",  # AggregationFormula
                float(posting.new_value) if isinstance(posting.new_value, Decimal) else posting.new_value,  # FinalPostedValue
                conf_str,  # Confidence
                "Y" if posting.had_conflict else "N",  # ConflictFlag
                posting.notes,  # Notes
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.BORDER
            row += 1

        return row + 1

    def _write_exceptions_section(
        self,
        ws,
        audit: RunAudit,
        start_row: int,
    ) -> int:
        """Write exceptions section."""
        row = self._write_section_header(ws, start_row, "EXCEPTIONS")

        # Headers
        headers = ["Category", "Severity", "Message", "Details"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        row += 1

        # Data
        for exc in audit.exceptions:
            ws.cell(row=row, column=1, value=exc.category)
            ws.cell(row=row, column=2, value=exc.severity)
            ws.cell(row=row, column=3, value=exc.message)
            ws.cell(row=row, column=4, value=str(exc.details) if exc.details else "")
            row += 1

        # Unmatched template items
        if audit.unmatched_template_items:
            for item in audit.unmatched_template_items:
                ws.cell(row=row, column=1, value="unmatched_template")
                ws.cell(row=row, column=2, value="warning")
                ws.cell(row=row, column=3, value=f"Template line item not matched: {item}")
                row += 1

        # Missing periods
        if audit.missing_periods:
            for period in audit.missing_periods:
                ws.cell(row=row, column=1, value="missing_period")
                ws.cell(row=row, column=2, value="warning")
                ws.cell(row=row, column=3, value=f"Period not found in PDF: {period}")
                row += 1

        if not audit.exceptions and not audit.unmatched_template_items and not audit.missing_periods:
            ws.cell(row=row, column=1, value="(No exceptions)")
            row += 1

        return row + 1

    def _write_reconciliation_section(
        self,
        ws,
        audit: RunAudit,
        start_row: int,
    ) -> int:
        """Write reconciliation results section."""
        row = self._write_section_header(ws, start_row, "RECONCILIATION CHECKS")

        if not audit.reconciliation:
            ws.cell(row=row, column=1, value="(No reconciliation checks performed)")
            return row + 2

        # Headers
        headers = ["Check", "Status", "Expected", "Actual", "Delta", "Delta %", "Severity", "Message"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        row += 1

        # Data
        for check in audit.reconciliation.checks:
            ws.cell(row=row, column=1, value=check.check_name)
            ws.cell(row=row, column=2, value="PASS" if check.is_valid else "FAIL")
            ws.cell(row=row, column=3, value=float(check.expected_value) if check.expected_value else "")
            ws.cell(row=row, column=4, value=float(check.actual_value) if check.actual_value else "")
            ws.cell(row=row, column=5, value=float(check.delta) if check.delta else "")
            ws.cell(row=row, column=6, value=f"{check.delta_percent:.2f}%" if check.delta_percent else "")
            ws.cell(row=row, column=7, value=check.severity)
            ws.cell(row=row, column=8, value=check.message)
            row += 1

        # Summary
        row += 1
        summary = f"Summary: {len([c for c in audit.reconciliation.checks if c.is_valid])}/{len(audit.reconciliation.checks)} checks passed"
        ws.cell(row=row, column=1, value=summary).font = Font(bold=True)

        return row + 2

    def _auto_fit_columns(self, ws) -> None:
        """Auto-fit column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width


def get_audit_sheet_generator() -> AuditSheetGenerator:
    """Get AuditSheetGenerator instance."""
    return AuditSheetGenerator()
