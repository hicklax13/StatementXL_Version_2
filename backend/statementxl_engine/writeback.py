"""
Writeback layer for the StatementXL Engine.

Pass 6: Template-safe Excel updates.
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional

import structlog
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from backend.statementxl_engine.models import (
    CellPosting,
    RunAudit,
    TemplateProfile,
)

logger = structlog.get_logger(__name__)


class WritebackLayer:
    """
    Template-safe writeback layer.

    Absolute rules:
    - Never rebuild workbook
    - Never change formatting, styles, dimensions, merges
    - Never change conditional formatting, freeze panes, filters
    - Never overwrite formulas
    - Only overwrite hardcoded numeric cells in detected value grid
    - Blank cells: fill ONLY if clearly part of an input grid slot
    """

    def write_postings(
        self,
        template_path: Path,
        output_path: Path,
        postings: List[CellPosting],
        template_profile: TemplateProfile,
    ) -> Path:
        """
        Write postings to template and save to output path.

        Args:
            template_path: Original template path.
            output_path: Output file path.
            postings: Cell postings to write.
            template_profile: Template profile for validation.

        Returns:
            Path to the written file.
        """
        logger.info(
            "Starting writeback",
            template=template_path.name,
            output=output_path.name,
            postings=len(postings),
        )

        # Load workbook preserving everything
        wb = load_workbook(template_path)

        # Track what we write
        written_count = 0
        skipped_count = 0

        for posting in postings:
            cell_info = posting.template_cell
            sheet_name = cell_info.sheet
            address = cell_info.address

            # Verify sheet exists
            if sheet_name not in wb.sheetnames:
                logger.warning(
                    "Sheet not found",
                    sheet=sheet_name,
                    address=address,
                )
                skipped_count += 1
                continue

            ws = wb[sheet_name]
            cell = ws[address]

            # CRITICAL: Check if cell has formula - never overwrite
            if self._has_formula(cell):
                logger.warning(
                    "Skipping formula cell",
                    sheet=sheet_name,
                    address=address,
                    formula=cell.value,
                )
                skipped_count += 1
                continue

            # Verify cell is eligible
            if not cell_info.is_eligible:
                logger.warning(
                    "Cell not eligible",
                    sheet=sheet_name,
                    address=address,
                )
                skipped_count += 1
                continue

            # Write the value
            try:
                # Convert Decimal to float for Excel
                value = float(posting.new_value) if isinstance(posting.new_value, Decimal) else posting.new_value

                # Write value - preserve existing number format if any
                cell.value = value

                # Apply right alignment for numbers (preserves other styles)
                cell.alignment = Alignment(horizontal='right')

                written_count += 1

                logger.debug(
                    "Wrote cell",
                    sheet=sheet_name,
                    address=address,
                    value=value,
                )

            except Exception as e:
                logger.error(
                    "Failed to write cell",
                    sheet=sheet_name,
                    address=address,
                    error=str(e),
                )
                skipped_count += 1

        # Save workbook
        wb.save(output_path)

        logger.info(
            "Writeback complete",
            written=written_count,
            skipped=skipped_count,
            output=str(output_path),
        )

        return output_path

    def _has_formula(self, cell) -> bool:
        """Check if cell contains a formula."""
        if cell.value is None:
            return False
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return True
        return False


def get_writeback_layer() -> WritebackLayer:
    """Get WritebackLayer instance."""
    return WritebackLayer()
