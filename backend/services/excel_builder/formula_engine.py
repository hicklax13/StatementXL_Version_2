"""
Formula Engine for injecting dynamic Excel formulas.

Replaces hardcoded values with SUM, DIFF, and other formulas
to create audit-ready, live-calculating Excel models.
"""

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import structlog
from openpyxl.utils import get_column_letter

from backend.services.excel_builder.template_parser import RowType, TemplateRow, TemplateStructure

logger = structlog.get_logger(__name__)


@dataclass
class FormulaSpec:
    """Specification for a formula to inject."""
    
    row: int
    formula_template: str  # e.g., "=SUM({start}:{end})"
    description: str


class FormulaEngine:
    """
    Injects dynamic Excel formulas into templates.
    
    Handles:
    - SUM formulas for subtotals
    - Difference formulas for profits/margins
    - Multi-column expansion
    """
    
    # Known formula patterns for Income Statement
    IS_FORMULA_PATTERNS = {
        "Total Revenue": {
            "type": "sum_children",
        },
        "Total Cost of Goods Sold": {
            "type": "sum_children",
        },
        "Gross Profit": {
            "type": "diff",
            "refs": ["Total Revenue", "Total Cost of Goods Sold"],
        },
        "Total Operating Expenses": {
            "type": "sum_children",
        },
        "Operating Income": {
            "type": "diff",
            "refs": ["Gross Profit", "Total Operating Expenses"],
        },
        "Total Other Income/(Expense)": {
            "type": "sum_children",
        },
        "Income Before Income Taxes": {
            "type": "diff",
            "refs": ["Operating Income", "Total Other Income/(Expense)"],
        },
        "Net Income": {
            "type": "diff",
            "refs": ["Income Before Income Taxes", "Provision for Income Taxes"],
        },
    }
    
    def __init__(self):
        """Initialize formula engine."""
        self._row_label_map: Dict[str, int] = {}
    
    def inject_formulas(
        self,
        structure: TemplateStructure,
        ws,  # openpyxl Worksheet
        period_columns: List[int],
    ) -> int:
        """
        Inject formulas into worksheet.
        
        Args:
            structure: Parsed template structure.
            ws: OpenPyXL worksheet to modify.
            period_columns: List of column indices with period data.
            
        Returns:
            Number of formulas injected.
        """
        logger.info(
            "Injecting formulas",
            rows=len(structure.rows),
            columns=len(period_columns),
        )
        
        # Build label to row mapping
        self._row_label_map = {
            row.label: row.row_num
            for row in structure.rows
            if row.label
        }
        
        formula_count = 0
        
        for row in structure.rows:
            if row.row_type not in (RowType.SUBTOTAL, RowType.TOTAL):
                continue
            
            # Get formula pattern
            pattern = self.IS_FORMULA_PATTERNS.get(row.label)
            
            if pattern:
                for col in period_columns:
                    formula = self._build_formula(row, pattern, col, structure)
                    if formula:
                        ws.cell(row=row.row_num, column=col, value=formula)
                        formula_count += 1
            
            elif row.row_type == RowType.SUBTOTAL and row.children:
                # Default SUM for subtotals with children
                for col in period_columns:
                    formula = self._build_sum_formula(row.children, col)
                    ws.cell(row=row.row_num, column=col, value=formula)
                    formula_count += 1
        
        logger.info("Formulas injected", count=formula_count)
        return formula_count
    
    def _build_formula(
        self,
        row: TemplateRow,
        pattern: Dict,
        col: int,
        structure: TemplateStructure,
    ) -> Optional[str]:
        """Build a formula based on pattern."""
        formula_type = pattern.get("type")
        col_letter = get_column_letter(col)
        
        if formula_type == "sum_children":
            if row.children:
                return self._build_sum_formula(row.children, col)
            return None
        
        elif formula_type == "diff":
            refs = pattern.get("refs", [])
            if len(refs) == 2:
                row1 = self._row_label_map.get(refs[0])
                row2 = self._row_label_map.get(refs[1])
                if row1 and row2:
                    return f"={col_letter}{row1}-{col_letter}{row2}"
            return None
        
        elif formula_type == "sum_refs":
            refs = pattern.get("refs", [])
            ref_rows = [self._row_label_map.get(r) for r in refs]
            ref_rows = [r for r in ref_rows if r]
            if ref_rows:
                refs_str = "+".join(f"{col_letter}{r}" for r in ref_rows)
                return f"={refs_str}"
            return None
        
        return None
    
    def _build_sum_formula(self, child_rows: List[int], col: int) -> str:
        """Build a SUM formula for child rows."""
        col_letter = get_column_letter(col)
        
        if not child_rows:
            return "=0"
        
        # Check if rows are contiguous
        sorted_rows = sorted(child_rows)
        if sorted_rows == list(range(sorted_rows[0], sorted_rows[-1] + 1)):
            # Contiguous range
            return f"=SUM({col_letter}{sorted_rows[0]}:{col_letter}{sorted_rows[-1]})"
        else:
            # Non-contiguous - list individual cells
            refs = ",".join(f"{col_letter}{r}" for r in sorted_rows)
            return f"=SUM({refs})"
    
    def get_formula_for_label(
        self,
        label: str,
        col: int,
        structure: TemplateStructure,
    ) -> Optional[str]:
        """Get the appropriate formula for a given label."""
        pattern = self.IS_FORMULA_PATTERNS.get(label)
        if not pattern:
            return None
        
        # Find the row for this label
        row = next(
            (r for r in structure.rows if r.label == label),
            None
        )
        
        if row:
            return self._build_formula(row, pattern, col, structure)
        
        return None


# Singleton instance
_engine_instance: Optional[FormulaEngine] = None


def get_formula_engine() -> FormulaEngine:
    """Get singleton FormulaEngine instance."""
    global _engine_instance
    if _engine_instance is None:
        _engine_instance = FormulaEngine()
    return _engine_instance
