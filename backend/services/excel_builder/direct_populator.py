"""
Direct Data Populator for generating Excel from PDF extractions.

Instead of matching extracted labels to fixed template labels,
this populator writes the extracted data directly to Excel,
preserving the original structure from the PDF.
"""

from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from backend.services.excel_builder.styles import STYLES, COLORWAYS, apply_colorway_to_style, StyleConfig

logger = structlog.get_logger(__name__)


class DirectPopulator:
    """
    Populates Excel directly from extracted PDF data.
    
    Key difference from ExcelBuilder:
    - Does NOT try to match labels to a template
    - Creates rows dynamically based on extracted data
    - Uses styling configuration for consistent formatting
    """
    
    # Column layout
    LABEL_COLUMN = 2  # B
    DATA_START_COLUMN = 3  # C
    
    def __init__(
        self,
        style: str = "basic",
        colorway: str = "green",
    ):
        """Initialize populator with styling options."""
        self.style_config = STYLES.get(style, STYLES["basic"])
        self.colorway_config = COLORWAYS.get(colorway, COLORWAYS["green"])
        self.applied_style = apply_colorway_to_style(self.style_config, self.colorway_config)
        
        self.workbook: Optional[Workbook] = None
        self.worksheet = None
    
    def populate(
        self,
        extracted_items: List[Dict[str, Any]],
        periods: List[int],
        company_name: Optional[str] = None,
        statement_title: str = "Income Statement",
    ) -> Workbook:
        """
        Generate Excel workbook from extracted items.
        
        Args:
            extracted_items: List of {label: str, period_year: value, ...}
            periods: List of period years [2025, 2024, ...]
            company_name: Optional company name for header
            statement_title: Title for the statement
            
        Returns:
            OpenPyXL Workbook with populated data
        """
        logger.info(
            "Direct population starting",
            items=len(extracted_items),
            periods=periods,
        )
        
        # Create workbook
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = statement_title
        
        # Sort periods (newest first)
        sorted_periods = sorted(periods, reverse=True)
        
        # Set column widths
        self.worksheet.column_dimensions['A'].width = 2  # Spacer
        self.worksheet.column_dimensions['B'].width = 45  # Labels
        for i, _ in enumerate(sorted_periods):
            col_letter = get_column_letter(self.DATA_START_COLUMN + i)
            self.worksheet.column_dimensions[col_letter].width = 15
        
        current_row = 1
        
        # Header: Company Name
        if company_name:
            cell = self.worksheet.cell(row=current_row, column=self.LABEL_COLUMN)
            cell.value = company_name
            cell.font = Font(bold=True, size=14)
            current_row += 1
        
        # Header: Statement Title
        cell = self.worksheet.cell(row=current_row, column=self.LABEL_COLUMN)
        cell.value = statement_title
        cell.font = Font(bold=True, size=12)
        current_row += 2  # Empty row after title
        
        # Period headers
        header_row = current_row
        self.worksheet.cell(row=header_row, column=self.LABEL_COLUMN, value="Line Item").font = Font(bold=True)
        for i, period in enumerate(sorted_periods):
            col = self.DATA_START_COLUMN + i
            cell = self.worksheet.cell(row=header_row, column=col)
            cell.value = period
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
        
        # Add bottom border to header
        for col in range(self.LABEL_COLUMN, self.DATA_START_COLUMN + len(sorted_periods)):
            cell = self.worksheet.cell(row=header_row, column=col)
            cell.border = Border(bottom=Side(style='medium'))
        
        current_row += 1
        data_start_row = current_row
        
        # Detect sections and organize items
        sections = self._organize_by_sections(extracted_items)
        
        # Populate data
        section_totals = {}
        items_in_section = {}
        
        for section_name, items in sections.items():
            # Section header
            if section_name:
                current_row += 1  # Add spacing before section
                cell = self.worksheet.cell(row=current_row, column=self.LABEL_COLUMN)
                cell.value = section_name
                cell.font = Font(bold=True)
                if self.applied_style.section_fill:
                    cell.fill = self.applied_style.section_fill
                current_row += 1
            
            section_items = []
            
            for item in items:
                label = item.get("label", "")
                is_total = self._is_total_row(label)
                
                # Write label
                cell = self.worksheet.cell(row=current_row, column=self.LABEL_COLUMN)
                cell.value = label
                
                if is_total:
                    cell.font = Font(bold=True)
                    cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
                else:
                    cell.font = Font(size=10)
                    section_items.append(current_row)
                
                # Write values for each period
                for i, period in enumerate(sorted_periods):
                    col = self.DATA_START_COLUMN + i
                    value = item.get(str(period), item.get(period, None))
                    
                    value_cell = self.worksheet.cell(row=current_row, column=col)
                    
                    if is_total:
                        # Generate SUM formula for totals (regardless of original value)
                        if section_items:
                            first_row = section_items[0]
                            last_row = section_items[-1]
                            col_letter = get_column_letter(col)
                            value_cell.value = f"=SUM({col_letter}{first_row}:{col_letter}{last_row})"
                        elif value is not None:
                            value_cell.value = float(value)
                        else:
                            value_cell.value = 0
                        value_cell.font = Font(bold=True)
                        value_cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
                        value_cell.number_format = '#,##0.00'
                        value_cell.alignment = Alignment(horizontal='right')
                    elif value is not None:
                        value_cell.value = float(value)
                        value_cell.number_format = '#,##0.00'
                        value_cell.alignment = Alignment(horizontal='right')
                
                current_row += 1
            
            # Track section total rows for overall total
            if section_name:
                items_in_section[section_name] = section_items
        
        # Add Overall Total if multiple sections
        if len(sections) > 1:
            current_row += 1
            cell = self.worksheet.cell(row=current_row, column=self.LABEL_COLUMN)
            cell.value = "Net Total"
            cell.font = Font(bold=True, size=11)
            cell.border = Border(top=Side(style='medium'), bottom=Side(style='double'))
            
            # Sum of all data rows
            for i, period in enumerate(sorted_periods):
                col = self.DATA_START_COLUMN + i
                col_letter = get_column_letter(col)
                all_data_rows = []
                for rows in items_in_section.values():
                    all_data_rows.extend(rows)
                if all_data_rows:
                    # Use explicit cell references
                    refs = "+".join(f"{col_letter}{r}" for r in all_data_rows)
                    value_cell = self.worksheet.cell(row=current_row, column=col)
                    value_cell.value = f"={refs}"
                    value_cell.font = Font(bold=True)
                    value_cell.number_format = '#,##0.00'
                    value_cell.alignment = Alignment(horizontal='right')
                    value_cell.border = Border(top=Side(style='medium'), bottom=Side(style='double'))
        
        logger.info(
            "Direct population complete",
            rows_written=current_row - data_start_row,
            sections=len(sections),
        )
        
        return self.workbook
    
    def _organize_by_sections(self, items: List[Dict]) -> Dict[str, List[Dict]]:
        """Organize items into sections based on labels."""
        sections = {}
        current_section = ""
        
        for item in items:
            label = item.get("label", "").strip()
            
            # Detect section headers (Income, Expenses, etc.)
            if self._is_section_header(label):
                current_section = label
                if current_section not in sections:
                    sections[current_section] = []
            elif label:
                if current_section not in sections:
                    sections[current_section] = []
                sections[current_section].append(item)
        
        return sections
    
    def _is_section_header(self, label: str) -> bool:
        """Check if label is a section header."""
        label_lower = label.lower()
        
        # Total/subtotal rows are NOT section headers
        total_keywords = ["total", "subtotal", "net", "gross", "overall", "sum", "balance"]
        for keyword in total_keywords:
            if keyword in label_lower:
                return False
        
        section_keywords = [
            "income", "revenue", "sales",
            "expenses", "expense", "costs", "cost of",
            "operating", "non-operating",
            "other income", "other expense",
        ]
        
        # Section headers are usually short and contain keywords
        if len(label.split()) <= 3:
            for keyword in section_keywords:
                if keyword in label_lower:
                    return True
        
        return False
    
    def _is_total_row(self, label: str) -> bool:
        """Check if label is a total row."""
        label_lower = label.lower()
        total_keywords = [
            "total", "subtotal", "net", "gross",
            "overall", "sum", "balance",
        ]
        
        for keyword in total_keywords:
            if keyword in label_lower:
                return True
        
        return False
    
    def save(self, output_path: Path) -> Path:
        """Save workbook to file."""
        if not self.workbook:
            raise ValueError("No workbook to save")
        
        self.workbook.save(output_path)
        logger.info("Workbook saved", path=str(output_path))
        return output_path


def populate_from_extraction(
    extracted_items: List[Dict],
    periods: List[int],
    output_path: Path,
    style: str = "basic",
    colorway: str = "green",
    company_name: Optional[str] = None,
    statement_title: str = "Income Statement",
) -> Path:
    """
    Convenience function to populate Excel from extracted data.
    
    Args:
        extracted_items: List of extracted line items
        periods: List of period years
        output_path: Where to save the file
        style: Style name
        colorway: Colorway name
        company_name: Optional company name
        statement_title: Title for the statement
        
    Returns:
        Path to saved file
    """
    populator = DirectPopulator(style=style, colorway=colorway)
    populator.populate(
        extracted_items=extracted_items,
        periods=periods,
        company_name=company_name,
        statement_title=statement_title,
    )
    return populator.save(output_path)
