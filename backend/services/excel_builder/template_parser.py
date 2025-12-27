"""
Template Parser for extracting hierarchy and structure from Excel templates.

Identifies sections, headers, line items, subtotals, and totals
to enable dynamic formula injection.
"""

from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import structlog
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

logger = structlog.get_logger(__name__)


class RowType(Enum):
    """Types of rows in a financial statement template."""
    
    HEADER = "header"  # Section header (e.g., "Income Statement", "Revenue")
    ITEM = "item"  # Input line item (e.g., "Products", "Services")
    SUBTOTAL = "subtotal"  # Sum of children (e.g., "Total Revenue")
    TOTAL = "total"  # Calculated total (e.g., "Gross Profit", "Net Income")
    SPACER = "spacer"  # Empty row for formatting


@dataclass
class TemplateRow:
    """Represents a row in the template with its structure."""
    
    row_num: int
    label: str
    row_type: RowType
    indent_level: int = 0  # 0 = section, 1 = category, 2 = item
    children: List[int] = field(default_factory=list)  # Child row numbers
    parent: Optional[int] = None  # Parent row number
    formula_type: Optional[str] = None  # "sum", "diff", etc.
    formula_refs: List[int] = field(default_factory=list)  # Rows referenced in formula


@dataclass
class TemplateStructure:
    """Complete parsed structure of a template."""
    
    rows: List[TemplateRow]
    label_column: int  # Column with labels (usually B = 2)
    data_start_column: int  # First data column (usually C = 3)
    header_row: int  # Row with period headers
    sections: Dict[str, Tuple[int, int]]  # {section_name: (start_row, end_row)}


class TemplateParser:
    """
    Parses Excel templates to extract hierarchical structure.
    
    Identifies:
    - Section headers (bold, all caps)
    - Line items (indented, regular text)
    - Subtotals (rows starting with "Total")
    - Calculated totals (Gross Profit, Operating Income, Net Income)
    """
    
    # Keywords indicating subtotal rows
    SUBTOTAL_KEYWORDS = ["total", "subtotal"]
    
    # Keywords indicating calculated totals (not sums)
    TOTAL_KEYWORDS = [
        "gross profit", "gross margin",
        "operating income", "operating profit",
        "income before", "earnings before",
        "net income", "net profit", "net earnings",
    ]
    
    # Keywords indicating section headers
    HEADER_KEYWORDS = [
        "income statement", "statement of operations",
        "revenue", "cost of goods sold", "cost of sales",
        "operating expenses", "other income", "other expense",
    ]
    
    def __init__(self, col_offset: int = 1):
        """
        Initialize parser.
        
        Args:
            col_offset: Column offset (1 means labels start in column B).
        """
        self.col_offset = col_offset
        self.label_column = 2  # Column B
        self.data_start_column = 3  # Column C
    
    def parse(self, template_path: Path) -> TemplateStructure:
        """
        Parse an Excel template file.
        
        Args:
            template_path: Path to .xlsx template file.
            
        Returns:
            TemplateStructure with parsed hierarchy.
        """
        logger.info("Parsing template", path=str(template_path))
        
        wb = load_workbook(template_path)
        ws = wb.active
        
        rows: List[TemplateRow] = []
        sections: Dict[str, Tuple[int, int]] = {}
        current_section: Optional[str] = None
        section_start: Optional[int] = None
        
        # Find header row (contains period like "2025")
        header_row = self._find_header_row(ws)
        
        # Parse each row
        for row_num in range(1, ws.max_row + 1):
            label_cell = ws.cell(row=row_num, column=self.label_column)
            label = str(label_cell.value or "").strip()
            
            if not label:
                # Empty row - spacer
                rows.append(TemplateRow(
                    row_num=row_num,
                    label="",
                    row_type=RowType.SPACER,
                ))
                continue
            
            # Determine row type
            row_type = self._classify_row(label, label_cell)
            indent_level = self._get_indent_level(label, label_cell)
            
            template_row = TemplateRow(
                row_num=row_num,
                label=label,
                row_type=row_type,
                indent_level=indent_level,
            )
            
            # Track sections
            if row_type == RowType.HEADER:
                if current_section and section_start:
                    sections[current_section] = (section_start, row_num - 1)
                current_section = label
                section_start = row_num
            
            rows.append(template_row)
        
        # Close last section
        if current_section and section_start:
            sections[current_section] = (section_start, ws.max_row)
        
        # Build parent-child relationships
        self._build_hierarchy(rows)
        
        # Determine formula types
        self._determine_formulas(rows)
        
        structure = TemplateStructure(
            rows=rows,
            label_column=self.label_column,
            data_start_column=self.data_start_column,
            header_row=header_row,
            sections=sections,
        )
        
        logger.info(
            "Template parsed",
            total_rows=len(rows),
            sections=len(sections),
            items=sum(1 for r in rows if r.row_type == RowType.ITEM),
            subtotals=sum(1 for r in rows if r.row_type == RowType.SUBTOTAL),
        )
        
        return structure
    
    def _find_header_row(self, ws: Worksheet) -> int:
        """Find the row containing period headers."""
        for row_num in range(1, min(10, ws.max_row + 1)):
            cell = ws.cell(row=row_num, column=self.data_start_column)
            value = str(cell.value or "")
            # Look for year or "Year(s) Ended"
            if value.isdigit() and len(value) == 4:
                return row_num
            if "year" in value.lower():
                return row_num
        return 4  # Default
    
    def _classify_row(self, label: str, cell: Cell) -> RowType:
        """Classify a row based on its label and formatting."""
        label_lower = label.lower()
        
        # Check for total keywords first
        for keyword in self.TOTAL_KEYWORDS:
            if keyword in label_lower:
                return RowType.TOTAL
        
        # Check for subtotal keywords
        for keyword in self.SUBTOTAL_KEYWORDS:
            if label_lower.startswith(keyword):
                return RowType.SUBTOTAL
        
        # Check for header keywords
        for keyword in self.HEADER_KEYWORDS:
            if keyword in label_lower:
                return RowType.HEADER
        
        # Check formatting (bold = header)
        if cell.font and cell.font.bold:
            return RowType.HEADER
        
        # Default to item
        return RowType.ITEM
    
    def _get_indent_level(self, label: str, cell: Cell) -> int:
        """Determine indent level of a row."""
        # Check cell alignment
        if cell.alignment and cell.alignment.indent:
            return int(cell.alignment.indent)
        
        # Check for leading spaces
        original = str(cell.value or "")
        stripped = original.lstrip()
        spaces = len(original) - len(stripped)
        
        return spaces // 2  # 2 spaces = 1 indent level
    
    def _build_hierarchy(self, rows: List[TemplateRow]) -> None:
        """Build parent-child relationships between rows."""
        # Find items that belong to subtotals
        for i, row in enumerate(rows):
            if row.row_type == RowType.SUBTOTAL:
                # Look backwards for items of the same indent or higher
                children = []
                for j in range(i - 1, -1, -1):
                    prev_row = rows[j]
                    if prev_row.row_type == RowType.SPACER:
                        continue
                    if prev_row.row_type == RowType.HEADER:
                        break  # Stop at section header
                    if prev_row.row_type == RowType.ITEM:
                        children.append(prev_row.row_num)
                        prev_row.parent = row.row_num
                    elif prev_row.row_type in (RowType.SUBTOTAL, RowType.TOTAL):
                        break  # Stop at another subtotal/total
                
                row.children = list(reversed(children))
    
    def _determine_formulas(self, rows: List[TemplateRow]) -> None:
        """Determine formula types for subtotal and total rows."""
        for row in rows:
            if row.row_type == RowType.SUBTOTAL:
                row.formula_type = "sum"
                row.formula_refs = row.children
            
            elif row.row_type == RowType.TOTAL:
                # Totals are typically differences
                row.formula_type = "diff"
                # Find the rows to subtract (context-dependent)
                # This will be refined based on specific patterns


# Singleton instance
_parser_instance: Optional[TemplateParser] = None


def get_template_parser() -> TemplateParser:
    """Get singleton TemplateParser instance."""
    global _parser_instance
    if _parser_instance is None:
        _parser_instance = TemplateParser()
    return _parser_instance
