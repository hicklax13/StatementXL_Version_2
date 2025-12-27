"""
Styles and Colorways for Excel template formatting.

Defines visual styling configurations for Basic, Corporate, and Professional
template styles with customizable color palettes.
"""

from dataclasses import dataclass, field
from typing import Dict, Optional

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


@dataclass
class StyleConfig:
    """Configuration for a template style."""
    
    name: str
    
    # Fonts
    title_font: Font
    header_font: Font
    section_font: Font
    item_font: Font
    total_font: Font
    
    # Fills (will be combined with colorway)
    title_fill: Optional[PatternFill] = None
    header_fill: Optional[PatternFill] = None
    section_fill: Optional[PatternFill] = None
    alt_row_fill: Optional[PatternFill] = None
    total_fill: Optional[PatternFill] = None
    
    # Borders
    header_border: Optional[Border] = None
    section_border: Optional[Border] = None
    total_border: Optional[Border] = None
    
    # Alignment
    label_alignment: Optional[Alignment] = None
    value_alignment: Optional[Alignment] = None


@dataclass
class Colorway:
    """Color palette for styling."""
    
    name: str
    primary: str  # Main accent color (hex without #)
    secondary: str  # Secondary color
    accent: str  # Highlight color
    text_on_primary: str = "FFFFFF"  # Text color when on primary bg
    text_dark: str = "000000"
    text_muted: str = "666666"
    border_color: str = "CCCCCC"
    alt_row_bg: str = "F5F5F5"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

STYLES: Dict[str, StyleConfig] = {
    "basic": StyleConfig(
        name="Basic",
        title_font=Font(bold=True, size=14),
        header_font=Font(bold=True, size=11),
        section_font=Font(bold=True, size=10),
        item_font=Font(size=10),
        total_font=Font(bold=True, size=10),
        header_border=Border(
            bottom=Side(style="thin", color="000000"),
        ),
        total_border=Border(
            top=Side(style="thin", color="000000"),
            bottom=Side(style="double", color="000000"),
        ),
        label_alignment=Alignment(horizontal="left", vertical="center"),
        value_alignment=Alignment(horizontal="right", vertical="center"),
    ),
    
    "corporate": StyleConfig(
        name="Corporate",
        title_font=Font(bold=True, size=16, color="FFFFFF"),
        header_font=Font(bold=True, size=11, color="FFFFFF"),
        section_font=Font(bold=True, size=10, color="FFFFFF"),
        item_font=Font(size=10),
        total_font=Font(bold=True, size=10),
        title_fill=PatternFill("solid", fgColor="2E7D32"),  # Will be replaced with colorway
        header_fill=PatternFill("solid", fgColor="2E7D32"),
        section_fill=PatternFill("solid", fgColor="4CAF50"),
        total_fill=PatternFill("solid", fgColor="E8F5E9"),
        header_border=Border(
            bottom=Side(style="medium", color="FFFFFF"),
        ),
        total_border=Border(
            top=Side(style="medium"),
            bottom=Side(style="double"),
        ),
        label_alignment=Alignment(horizontal="left", vertical="center"),
        value_alignment=Alignment(horizontal="right", vertical="center"),
    ),
    
    "professional": StyleConfig(
        name="Professional",
        title_font=Font(bold=True, size=14),
        header_font=Font(bold=True, size=11),
        section_font=Font(bold=True, size=10, italic=True),
        item_font=Font(size=10),
        total_font=Font(bold=True, size=10),
        alt_row_fill=PatternFill("solid", fgColor="F5F5F5"),
        header_border=Border(
            bottom=Side(style="medium"),
        ),
        section_border=Border(
            bottom=Side(style="thin", color="CCCCCC"),
        ),
        total_border=Border(
            top=Side(style="thin"),
            bottom=Side(style="medium"),
        ),
        label_alignment=Alignment(horizontal="left", vertical="center", indent=1),
        value_alignment=Alignment(horizontal="right", vertical="center"),
    ),
}


# =============================================================================
# COLORWAY DEFINITIONS
# =============================================================================

COLORWAYS: Dict[str, Colorway] = {
    "green": Colorway(
        name="Green",
        primary="2E7D32",  # Dark green
        secondary="4CAF50",  # Medium green
        accent="81C784",  # Light green
        alt_row_bg="E8F5E9",  # Very light green
    ),
    
    "blue": Colorway(
        name="Blue",
        primary="1565C0",  # Dark blue
        secondary="1E88E5",  # Medium blue
        accent="42A5F5",  # Light blue
        alt_row_bg="E3F2FD",  # Very light blue
    ),
    
    "navy": Colorway(
        name="Navy",
        primary="1A237E",  # Navy
        secondary="303F9F",  # Indigo
        accent="5C6BC0",  # Light indigo
        alt_row_bg="E8EAF6",  # Very light indigo
    ),
    
    "slate": Colorway(
        name="Slate",
        primary="37474F",  # Blue grey
        secondary="546E7A",  # Medium grey
        accent="78909C",  # Light grey
        alt_row_bg="ECEFF1",  # Very light grey
    ),
    
    "teal": Colorway(
        name="Teal",
        primary="00695C",  # Dark teal
        secondary="00897B",  # Medium teal
        accent="4DB6AC",  # Light teal
        alt_row_bg="E0F2F1",  # Very light teal
    ),
    
    "burgundy": Colorway(
        name="Burgundy",
        primary="7B1FA2",  # Purple
        secondary="9C27B0",  # Medium purple
        accent="BA68C8",  # Light purple
        alt_row_bg="F3E5F5",  # Very light purple
    ),
    
    "charcoal": Colorway(
        name="Charcoal",
        primary="212121",  # Near black
        secondary="424242",  # Dark grey
        accent="757575",  # Medium grey
        alt_row_bg="FAFAFA",  # Near white
    ),
}


def apply_colorway_to_style(style: StyleConfig, colorway: Colorway) -> StyleConfig:
    """
    Apply a colorway to a style configuration.
    
    Creates a new StyleConfig with colors updated to match the colorway.
    """
    # Create new fill objects with colorway colors
    new_title_fill = None
    new_header_fill = None
    new_section_fill = None
    new_alt_row_fill = None
    new_total_fill = None
    
    if style.title_fill:
        new_title_fill = PatternFill("solid", fgColor=colorway.primary)
    
    if style.header_fill:
        new_header_fill = PatternFill("solid", fgColor=colorway.primary)
    
    if style.section_fill:
        new_section_fill = PatternFill("solid", fgColor=colorway.secondary)
    
    if style.alt_row_fill:
        new_alt_row_fill = PatternFill("solid", fgColor=colorway.alt_row_bg)
    
    if style.total_fill:
        new_total_fill = PatternFill("solid", fgColor=colorway.alt_row_bg)
    
    # Create updated fonts for colored backgrounds
    new_title_font = Font(
        bold=style.title_font.bold,
        size=style.title_font.size,
        color=colorway.text_on_primary if style.title_fill else style.title_font.color,
    )
    
    new_header_font = Font(
        bold=style.header_font.bold,
        size=style.header_font.size,
        color=colorway.text_on_primary if style.header_fill else style.header_font.color,
    )
    
    new_section_font = Font(
        bold=style.section_font.bold,
        size=style.section_font.size,
        italic=style.section_font.italic,
        color=colorway.text_on_primary if style.section_fill else style.section_font.color,
    )
    
    return StyleConfig(
        name=f"{style.name} - {colorway.name}",
        title_font=new_title_font,
        header_font=new_header_font,
        section_font=new_section_font,
        item_font=style.item_font,
        total_font=style.total_font,
        title_fill=new_title_fill,
        header_fill=new_header_fill,
        section_fill=new_section_fill,
        alt_row_fill=new_alt_row_fill,
        total_fill=new_total_fill,
        header_border=style.header_border,
        section_border=style.section_border,
        total_border=style.total_border,
        label_alignment=style.label_alignment,
        value_alignment=style.value_alignment,
    )
