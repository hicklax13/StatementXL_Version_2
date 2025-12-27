"""
Main ExcelBuilder class for generating formatted Excel statements.

Orchestrates template parsing, data population, formula injection,
and styling to produce audit-ready Excel financial statements.
"""

import shutil
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Literal, Optional

import structlog
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.services.excel_builder.template_parser import (
    RowType,
    TemplateParser,
    TemplateStructure,
    get_template_parser,
)
from backend.services.excel_builder.formula_engine import (
    FormulaEngine,
    get_formula_engine,
)
from backend.services.excel_builder.aggregator import (
    AggregatedItem,
    GAAPAggregator,
    get_gaap_aggregator,
)
from backend.services.excel_builder.styles import (
    STYLES,
    COLORWAYS,
    StyleConfig,
    Colorway,
    apply_colorway_to_style,
)

logger = structlog.get_logger(__name__)


StatementType = Literal["income_statement", "balance_sheet", "cash_flow"]
StyleType = Literal["basic", "corporate", "professional"]


class ExcelBuilder:
    """
    Generates formatted Excel statements from extracted PDF data.
    
    Main orchestrator that:
    1. Loads and parses template structure
    2. Expands columns for multiple periods
    3. Populates data from aggregated extractions
    4. Injects dynamic formulas
    5. Applies styling and colorway
    """
    
    # Column offset (Column A is spacer)
    COL_OFFSET = 1
    
    # Label column (B = 2)
    LABEL_COLUMN = 2
    
    # First data column (C = 3)
    DATA_START_COLUMN = 3
    
    def __init__(
        self,
        template_path: Path,
        style: StyleType = "basic",
        colorway: str = "green",
    ):
        """
        Initialize ExcelBuilder.
        
        Args:
            template_path: Path to the Excel template file.
            style: Style variant (basic, corporate, professional).
            colorway: Color palette name.
        """
        self.template_path = template_path
        self.style_name = style
        self.colorway_name = colorway
        
        # Get style and colorway
        self.style: StyleConfig = STYLES.get(style, STYLES["basic"])
        self.colorway: Colorway = COLORWAYS.get(colorway, COLORWAYS["green"])
        
        # Apply colorway to style
        self.applied_style = apply_colorway_to_style(self.style, self.colorway)
        
        # Initialize components
        self.parser: TemplateParser = get_template_parser()
        self.formula_engine: FormulaEngine = get_formula_engine()
        self.aggregator: GAAPAggregator = get_gaap_aggregator()
        
        # State
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.structure: Optional[TemplateStructure] = None
        self.period_columns: List[int] = []
    
    def build(
        self,
        aggregated_data: Dict[str, AggregatedItem],
        periods: List[int],
        company_name: Optional[str] = None,
    ) -> Workbook:
        """
        Build the Excel workbook from aggregated data.
        
        Args:
            aggregated_data: Dict mapping template labels to aggregated values.
            periods: List of years/periods (e.g., [2025, 2024, 2023]).
            company_name: Optional company name for the header.
            
        Returns:
            OpenPyXL Workbook object.
        """
        logger.info(
            "Building Excel statement",
            template=str(self.template_path),
            style=self.style_name,
            colorway=self.colorway_name,
            periods=periods,
        )
        
        # Load template
        self.workbook = load_workbook(self.template_path)
        self.worksheet = self.workbook.active
        
        # Parse structure
        self.structure = self.parser.parse(self.template_path)
        
        # Sort periods (newest first)
        sorted_periods = sorted(periods, reverse=True)
        
        # Expand columns for periods
        self._expand_columns(sorted_periods)
        
        # Set period headers
        self._set_period_headers(sorted_periods)
        
        # Set company name if provided
        if company_name:
            self._set_company_name(company_name)
        
        # Populate data
        self._populate_data(aggregated_data, sorted_periods)
        
        # Inject formulas
        formula_count = self.formula_engine.inject_formulas(
            self.structure,
            self.worksheet,
            self.period_columns,
        )
        
        # Apply styling
        self._apply_styling()
        
        logger.info(
            "Excel statement built",
            formulas_injected=formula_count,
            rows_populated=len(aggregated_data),
        )
        
        return self.workbook
    
    def _expand_columns(self, periods: List[int]) -> None:
        """Expand template to accommodate multiple periods."""
        num_periods = len(periods)
        
        # Determine column indices
        self.period_columns = [
            self.DATA_START_COLUMN + i
            for i in range(num_periods)
        ]
        
        if num_periods <= 1:
            return
        
        # Copy column C formatting to D, E, etc.
        base_col = self.DATA_START_COLUMN
        
        for row in range(1, self.worksheet.max_row + 1):
            base_cell = self.worksheet.cell(row=row, column=base_col)
            
            for i in range(1, num_periods):
                new_col = base_col + i
                new_cell = self.worksheet.cell(row=row, column=new_col)
                
                # Copy formatting
                if base_cell.has_style:
                    new_cell.font = base_cell.font.copy()
                    new_cell.fill = base_cell.fill.copy()
                    new_cell.border = base_cell.border.copy()
                    new_cell.alignment = base_cell.alignment.copy()
                    new_cell.number_format = base_cell.number_format
        
        # Copy column width
        base_width = self.worksheet.column_dimensions[get_column_letter(base_col)].width
        for col in self.period_columns[1:]:
            self.worksheet.column_dimensions[get_column_letter(col)].width = base_width
        
        logger.debug("Columns expanded", count=num_periods)
    
    def _set_period_headers(self, periods: List[int]) -> None:
        """Set year headers in the header row."""
        header_row = self.structure.header_row if self.structure else 4
        
        for i, year in enumerate(periods):
            col = self.period_columns[i]
            self.worksheet.cell(row=header_row, column=col, value=year)
    
    def _set_company_name(self, company_name: str) -> None:
        """Set the company name in the template."""
        # Look for "Company Name" placeholder in first few rows
        for row in range(1, 5):
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value and "company" in str(cell.value).lower():
                cell.value = company_name
                return
        
        # Default to cell A1
        self.worksheet.cell(row=1, column=1, value=company_name)
    
    def _populate_data(
        self,
        aggregated_data: Dict[str, AggregatedItem],
        periods: List[int],
    ) -> None:
        """Populate template with aggregated data values."""
        if not self.structure:
            return
        
        for row in self.structure.rows:
            if row.row_type != RowType.ITEM:
                continue
            
            # Find matching aggregated item
            agg_item = aggregated_data.get(row.label)
            if not agg_item:
                # Try partial match
                for label, item in aggregated_data.items():
                    if row.label.lower() in label.lower() or label.lower() in row.label.lower():
                        agg_item = item
                        break
            
            if agg_item:
                for i, year in enumerate(periods):
                    col = self.period_columns[i]
                    value = agg_item.values.get(year, Decimal("0"))
                    self.worksheet.cell(row=row.row_num, column=col, value=float(value))
    
    def _apply_styling(self) -> None:
        """Apply style and colorway to the worksheet."""
        if not self.structure:
            return
        
        for row in self.structure.rows:
            if row.row_type == RowType.SPACER:
                continue
            
            # Get appropriate styling
            if row.row_type == RowType.HEADER:
                font = self.applied_style.section_font
                fill = self.applied_style.section_fill
                border = self.applied_style.section_border
            elif row.row_type in (RowType.SUBTOTAL, RowType.TOTAL):
                font = self.applied_style.total_font
                fill = self.applied_style.total_fill
                border = self.applied_style.total_border
            else:
                font = self.applied_style.item_font
                fill = None
                border = None
            
            # Apply to label cell
            label_cell = self.worksheet.cell(row=row.row_num, column=self.LABEL_COLUMN)
            if font:
                label_cell.font = font
            if fill:
                label_cell.fill = fill
            if border:
                label_cell.border = border
            if self.applied_style.label_alignment:
                label_cell.alignment = self.applied_style.label_alignment
            
            # Apply to data cells
            for col in self.period_columns:
                data_cell = self.worksheet.cell(row=row.row_num, column=col)
                if font:
                    data_cell.font = font
                if fill:
                    data_cell.fill = fill
                if border:
                    data_cell.border = border
                if self.applied_style.value_alignment:
                    data_cell.alignment = self.applied_style.value_alignment
    
    def save(self, output_path: Path) -> Path:
        """
        Save the workbook to a file.
        
        Args:
            output_path: Path to save the Excel file.
            
        Returns:
            Path to the saved file.
        """
        if not self.workbook:
            raise ValueError("No workbook to save. Call build() first.")
        
        self.workbook.save(output_path)
        logger.info("Workbook saved", path=str(output_path))
        
        return output_path
    
    @classmethod
    def from_extracted_data(
        cls,
        template_path: Path,
        extracted_items: List[Dict],
        periods: List[int],
        style: StyleType = "basic",
        colorway: str = "green",
        company_name: Optional[str] = None,
    ) -> "ExcelBuilder":
        """
        Factory method to create builder and process extracted data.
        
        Args:
            template_path: Path to the Excel template.
            extracted_items: Raw extracted line items from PDF.
            periods: List of periods in the data.
            style: Style variant.
            colorway: Color palette name.
            company_name: Optional company name.
            
        Returns:
            ExcelBuilder instance with built workbook.
        """
        builder = cls(template_path, style, colorway)
        
        # Aggregate extracted items
        aggregated = builder.aggregator.aggregate(extracted_items, periods)
        
        # Build workbook
        builder.build(aggregated, periods, company_name)
        
        return builder


def get_excel_builder(
    template_path: Path,
    style: StyleType = "basic",
    colorway: str = "green",
) -> ExcelBuilder:
    """Create a new ExcelBuilder instance."""
    return ExcelBuilder(template_path, style, colorway)
