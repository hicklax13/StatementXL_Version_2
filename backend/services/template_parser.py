"""
Template Parser Service

Parses Excel template files to extract:
- Row structure and positions
- Section headers and their locations
- Formula cells vs value cells
- Formatting rules
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

import structlog
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

logger = structlog.get_logger(__name__)


@dataclass
class TemplateRow:
    """Represents a row in the template."""
    
    row_number: int
    label: str
    is_section_header: bool = False
    is_total: bool = False
    is_calculated: bool = False
    has_formula: bool = False
    formula: Optional[str] = None
    indent_level: int = 0
    has_currency_symbol: bool = False
    fill_color: Optional[str] = None


@dataclass
class TemplateColumn:
    """Represents a data column in the template."""
    
    column_number: int
    column_letter: str
    header: Optional[str] = None
    is_period: bool = False
    period_year: Optional[int] = None


@dataclass
class TemplateStructure:
    """Complete parsed template structure."""
    
    file_path: Path
    worksheet_name: str
    rows: Dict[str, TemplateRow] = field(default_factory=dict)
    columns: List[TemplateColumn] = field(default_factory=list)
    label_column: int = 2  # Column B by default
    data_start_column: int = 3  # Column C by default
    section_headers: List[str] = field(default_factory=list)
    total_rows: List[str] = field(default_factory=list)
    calculated_rows: List[str] = field(default_factory=list)


class TemplateParser:
    """
    Parses Excel template files to understand their structure.
    
    This enables the populator to write values to the correct cells
    while preserving the template's formatting and formulas.
    """
    
    # Default template paths - use Excel Templates folder
    TEMPLATE_DIR = Path("Excel Templates")
    
    TEMPLATES = {
        "income_statement": {
            "basic": TEMPLATE_DIR / "income_statement" / "Basic_Style_IS" / "StatementXL_Income_Statement_Template_Basic.xlsx",
            "corporate": TEMPLATE_DIR / "income_statement" / "Corporate_Style_IS" / "StatementXL_Income_Statement_Template_Corporate.xlsx",
            "professional": TEMPLATE_DIR / "income_statement" / "Professional_Style_IS" / "StatementXL_Income_Statement_Template_Professional.xlsx",
        },
        "balance_sheet": {
            "basic": TEMPLATE_DIR / "balance_sheet" / "Basic_Style" / "StatementXL_Balance_Sheet_Template_Basic.xlsx",
            "corporate": TEMPLATE_DIR / "balance_sheet" / "Corporate_Style" / "StatementXL_Balance_Sheet_Template_Corporate.xlsx",
            "professional": TEMPLATE_DIR / "balance_sheet" / "Professional_Style" / "StatementXL_Balance_Sheet_Template_Professional.xlsx",
        },
        "cash_flow": {
            "basic": TEMPLATE_DIR / "cash_flow" / "basic.xlsx",
            "corporate": TEMPLATE_DIR / "cash_flow" / "corporate.xlsx",
            "professional": TEMPLATE_DIR / "cash_flow" / "professional.xlsx",
        },
    }
    
    # Section header keywords
    SECTION_KEYWORDS = [
        "revenue", "cost of goods sold", "operating expenses",
        "other income", "assets", "liabilities", "equity",
        "operating activities", "investing activities", "financing activities"
    ]
    
    # Total row keywords
    TOTAL_KEYWORDS = ["total", "subtotal"]
    
    # Calculated row keywords
    CALCULATED_KEYWORDS = [
        "gross profit", "operating income", "net income",
        "income before", "ebit", "ebitda"
    ]
    
    def __init__(self):
        """Initialize the template parser."""
        self._cache: Dict[str, TemplateStructure] = {}
    
    def parse(
        self,
        statement_type: str = "income_statement",
        style: str = "basic"
    ) -> TemplateStructure:
        """
        Parse a template file and return its structure.
        
        Args:
            statement_type: Type of statement (income_statement, balance_sheet, cash_flow)
            style: Template style (basic, corporate, professional)
            
        Returns:
            TemplateStructure with parsed information
        """
        cache_key = f"{statement_type}_{style}"
        
        if cache_key in self._cache:
            return self._cache[cache_key]
        
        # Get template path
        template_path = self._get_template_path(statement_type, style)
        
        if not template_path.exists():
            logger.warning(
                "Template not found, using basic",
                path=str(template_path),
                statement_type=statement_type,
                style=style,
            )
            # Fall back to basic if style not found
            template_path = self._get_template_path(statement_type, "basic")
        
        logger.info("Parsing template", path=str(template_path))
        
        # Load workbook
        wb = load_workbook(template_path, data_only=False)
        ws = wb.active
        
        # Parse structure
        structure = TemplateStructure(
            file_path=template_path,
            worksheet_name=ws.title,
        )
        
        # Parse rows
        self._parse_rows(ws, structure)
        
        # Parse columns
        self._parse_columns(ws, structure)
        
        # Cache and return
        self._cache[cache_key] = structure
        
        logger.info(
            "Template parsed",
            rows=len(structure.rows),
            sections=len(structure.section_headers),
            totals=len(structure.total_rows),
        )
        
        return structure
    
    def _get_template_path(self, statement_type: str, style: str) -> Path:
        """Get the template file path."""
        if statement_type in self.TEMPLATES and style in self.TEMPLATES[statement_type]:
            return self.TEMPLATES[statement_type][style]
        
        # Default to income statement basic
        return self.TEMPLATES["income_statement"]["basic"]
    
    def _parse_rows(self, ws: Worksheet, structure: TemplateStructure) -> None:
        """Parse all rows in the template."""
        
        for row_num in range(1, ws.max_row + 1):
            # Get label from column B (column 2)
            label_cell = ws.cell(row=row_num, column=structure.label_column)
            label = str(label_cell.value).strip() if label_cell.value else ""
            
            if not label:
                continue
            
            # Get data cell from column C (column 3)
            data_cell = ws.cell(row=row_num, column=structure.data_start_column)
            
            # Check for currency symbol in column B or between
            currency_cell = ws.cell(row=row_num, column=structure.label_column + 1)
            has_currency = currency_cell.value == "$" if currency_cell.value else False
            
            # Determine row type
            is_section = self._is_section_header(label, label_cell)
            is_total = self._is_total_row(label)
            is_calculated = self._is_calculated_row(label)
            
            # Check for formula
            has_formula = False
            formula_str = None
            if data_cell.value and isinstance(data_cell.value, str) and data_cell.value.startswith("="):
                has_formula = True
                formula_str = data_cell.value
            
            # Determine indent level from cell formatting or position
            indent_level = self._get_indent_level(label_cell)
            
            # Get fill color if any
            fill_color = self._get_fill_color(label_cell)
            
            # Create row object
            template_row = TemplateRow(
                row_number=row_num,
                label=label,
                is_section_header=is_section,
                is_total=is_total,
                is_calculated=is_calculated,
                has_formula=has_formula,
                formula=formula_str,
                indent_level=indent_level,
                has_currency_symbol=has_currency,
                fill_color=fill_color,
            )
            
            # Store in structure - ONLY if not already present (keep first occurrence)
            # This ensures Products (row 7 under Revenue) is kept, not overwritten by Products (row 13 under COGS)
            if label not in structure.rows:
                structure.rows[label] = template_row
            
            # Track categories
            if is_section:
                structure.section_headers.append(label)
            if is_total:
                structure.total_rows.append(label)
            if is_calculated:
                structure.calculated_rows.append(label)
    
    def _parse_columns(self, ws: Worksheet, structure: TemplateStructure) -> None:
        """Parse column headers to identify period columns."""
        
        from openpyxl.utils import get_column_letter
        
        # Check header row (usually row 4 or 5)
        for header_row in [4, 5]:
            for col_num in range(structure.data_start_column, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col_num)
                header_value = str(cell.value) if cell.value else ""
                
                # Try to detect if this is a year
                is_period = False
                period_year = None
                
                if header_value:
                    # Check for year pattern
                    if header_value.isdigit() and len(header_value) == 4:
                        is_period = True
                        period_year = int(header_value)
                    elif "year" in header_value.lower():
                        is_period = True
                
                if header_value or is_period:
                    structure.columns.append(
                        TemplateColumn(
                            column_number=col_num,
                            column_letter=get_column_letter(col_num),
                            header=header_value,
                            is_period=is_period,
                            period_year=period_year,
                        )
                    )
    
    def _is_section_header(self, label: str, cell: Cell) -> bool:
        """Check if a row is a section header."""
        label_lower = label.lower()
        
        # Check keywords
        for keyword in self.SECTION_KEYWORDS:
            if keyword in label_lower:
                return True
        
        # Check for background fill (section headers often have colored backgrounds)
        if cell.fill and cell.fill.patternType == "solid":
            return True
        
        return False
    
    def _is_total_row(self, label: str) -> bool:
        """Check if a row is a total row."""
        label_lower = label.lower()
        return any(keyword in label_lower for keyword in self.TOTAL_KEYWORDS)
    
    def _is_calculated_row(self, label: str) -> bool:
        """Check if a row is a calculated row (like Gross Profit)."""
        label_lower = label.lower()
        return any(keyword in label_lower for keyword in self.CALCULATED_KEYWORDS)
    
    def _get_indent_level(self, cell: Cell) -> int:
        """Determine the indent level of a cell."""
        # Check alignment indent
        if cell.alignment and cell.alignment.indent:
            return cell.alignment.indent
        
        # Check for leading spaces in value
        if cell.value and isinstance(cell.value, str):
            stripped = cell.value.lstrip()
            return len(cell.value) - len(stripped)
        
        return 0
    
    def _get_fill_color(self, cell: Cell) -> Optional[str]:
        """Get the fill color of a cell if any."""
        if cell.fill and cell.fill.patternType == "solid":
            if cell.fill.fgColor and cell.fill.fgColor.rgb:
                return str(cell.fill.fgColor.rgb)
        return None
    
    def get_row_for_label(
        self,
        structure: TemplateStructure,
        label: str
    ) -> Optional[TemplateRow]:
        """
        Find the template row that matches a label.
        
        Supports exact match and fuzzy matching for common variations.
        """
        # Exact match
        if label in structure.rows:
            return structure.rows[label]
        
        # Case-insensitive match
        label_lower = label.lower()
        for template_label, row in structure.rows.items():
            if template_label.lower() == label_lower:
                return row
        
        # Partial match
        for template_label, row in structure.rows.items():
            if label_lower in template_label.lower() or template_label.lower() in label_lower:
                return row
        
        return None


# Singleton instance
_parser_instance: Optional[TemplateParser] = None


def get_template_parser() -> TemplateParser:
    """Get singleton TemplateParser instance."""
    global _parser_instance
    if _parser_instance is None:
        _parser_instance = TemplateParser()
    return _parser_instance
