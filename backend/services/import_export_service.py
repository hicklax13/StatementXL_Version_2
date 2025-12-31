"""
Data Import/Export Service.

Provides comprehensive data import and export capabilities with multiple formats.
"""
import csv
import io
import json
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any, BinaryIO, Tuple
from dataclasses import dataclass
from enum import Enum

import structlog
from sqlalchemy.orm import Session

from backend.config import get_settings
from backend.models.document import Document
from backend.models.template import Template

logger = structlog.get_logger(__name__)
settings = get_settings()


class ExportFormat(str, Enum):
    """Supported export formats."""
    JSON = "json"
    CSV = "csv"
    XLSX = "xlsx"
    XML = "xml"
    ZIP = "zip"


class ImportFormat(str, Enum):
    """Supported import formats."""
    JSON = "json"
    CSV = "csv"
    XLSX = "xlsx"
    XML = "xml"


@dataclass
class ExportOptions:
    """Options for data export."""
    format: ExportFormat = ExportFormat.JSON
    include_metadata: bool = True
    include_relationships: bool = False
    fields: Optional[List[str]] = None  # None = all fields
    pretty_print: bool = False
    compress: bool = False
    password: Optional[str] = None  # For encrypted exports
    date_format: str = "%Y-%m-%d %H:%M:%S"


@dataclass
class ImportResult:
    """Result of import operation."""
    success: bool
    total_records: int
    imported: int
    skipped: int
    errors: List[Dict[str, Any]]
    warnings: List[str]


class ImportExportService:
    """
    Service for importing and exporting data in multiple formats.
    """

    def __init__(self, db: Session):
        self.db = db

    # ==================== Export Methods ====================

    def export_documents(
        self,
        user_id: uuid.UUID,
        options: ExportOptions = None,
        document_ids: Optional[List[uuid.UUID]] = None,
    ) -> Tuple[bytes, str]:
        """
        Export documents to specified format.

        Args:
            user_id: User performing export
            options: Export options
            document_ids: Specific documents to export (None = all)

        Returns:
            Tuple of (file_content, filename)
        """
        options = options or ExportOptions()

        # Query documents
        query = self.db.query(Document).filter(Document.user_id == user_id)
        if document_ids:
            query = query.filter(Document.id.in_(document_ids))

        documents = query.all()

        # Serialize documents
        data = self._serialize_documents(documents, options)

        # Export based on format
        if options.format == ExportFormat.JSON:
            content = self._export_json(data, options)
            filename = f"documents_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
        elif options.format == ExportFormat.CSV:
            content = self._export_csv(data, options)
            filename = f"documents_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        elif options.format == ExportFormat.XLSX:
            content = self._export_xlsx(data, options)
            filename = f"documents_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif options.format == ExportFormat.XML:
            content = self._export_xml(data, options)
            filename = f"documents_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xml"
        else:
            raise ValueError(f"Unsupported format: {options.format}")

        # Compress if requested
        if options.compress:
            content = self._compress(content, filename)
            filename = filename.rsplit(".", 1)[0] + ".zip"

        return content, filename

    def export_templates(
        self,
        user_id: uuid.UUID,
        options: ExportOptions = None,
        template_ids: Optional[List[uuid.UUID]] = None,
    ) -> Tuple[bytes, str]:
        """Export templates to specified format."""
        options = options or ExportOptions()

        # Query templates
        query = self.db.query(Template)
        if template_ids:
            query = query.filter(Template.id.in_(template_ids))

        templates = query.all()

        # Serialize templates
        data = self._serialize_templates(templates, options)

        # Export based on format
        if options.format == ExportFormat.JSON:
            content = self._export_json(data, options)
            filename = f"templates_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
        elif options.format == ExportFormat.CSV:
            content = self._export_csv(data, options)
            filename = f"templates_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            content = self._export_json(data, options)
            filename = f"templates_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"

        return content, filename

    def export_full_backup(
        self,
        user_id: uuid.UUID,
        organization_id: Optional[uuid.UUID] = None,
    ) -> Tuple[bytes, str]:
        """
        Export full data backup as ZIP archive.

        Includes:
        - Documents
        - Templates
        - Mappings
        - Settings
        """
        buffer = io.BytesIO()
        filename = f"statementxl_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"

        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Export documents
            docs_content, _ = self.export_documents(
                user_id,
                ExportOptions(format=ExportFormat.JSON, include_relationships=True),
            )
            zf.writestr("documents.json", docs_content)

            # Export templates
            templates_content, _ = self.export_templates(
                user_id,
                ExportOptions(format=ExportFormat.JSON, include_relationships=True),
            )
            zf.writestr("templates.json", templates_content)

            # Add manifest
            manifest = {
                "version": "2.0.0",
                "created_at": datetime.utcnow().isoformat(),
                "user_id": str(user_id),
                "organization_id": str(organization_id) if organization_id else None,
                "contents": ["documents.json", "templates.json"],
            }
            zf.writestr("manifest.json", json.dumps(manifest, indent=2))

        buffer.seek(0)
        return buffer.read(), filename

    def _serialize_documents(
        self,
        documents: List[Document],
        options: ExportOptions,
    ) -> List[Dict[str, Any]]:
        """Serialize documents to dictionaries."""
        data = []
        for doc in documents:
            item = {
                "id": str(doc.id),
                "filename": doc.filename,
                "status": doc.status.value if hasattr(doc.status, 'value') else str(doc.status),
                "page_count": doc.page_count,
                "created_at": doc.created_at.strftime(options.date_format) if doc.created_at else None,
                "updated_at": doc.updated_at.strftime(options.date_format) if doc.updated_at else None,
            }

            if options.include_metadata:
                item["file_path"] = doc.file_path
                item["error_message"] = doc.error_message

            if options.fields:
                item = {k: v for k, v in item.items() if k in options.fields}

            data.append(item)

        return data

    def _serialize_templates(
        self,
        templates: List[Template],
        options: ExportOptions,
    ) -> List[Dict[str, Any]]:
        """Serialize templates to dictionaries."""
        data = []
        for template in templates:
            item = {
                "id": str(template.id),
                "filename": template.filename,
                "status": template.status.value if hasattr(template.status, 'value') else str(template.status),
                "sheet_count": template.sheet_count,
                "created_at": template.created_at.strftime(options.date_format) if template.created_at else None,
            }

            if options.include_metadata and template.extra_data:
                item["extra_data"] = template.extra_data

            if options.fields:
                item = {k: v for k, v in item.items() if k in options.fields}

            data.append(item)

        return data

    def _export_json(self, data: List[Dict], options: ExportOptions) -> bytes:
        """Export data as JSON."""
        indent = 2 if options.pretty_print else None
        return json.dumps(data, indent=indent, default=str).encode('utf-8')

    def _export_csv(self, data: List[Dict], options: ExportOptions) -> bytes:
        """Export data as CSV."""
        if not data:
            return b""

        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        return buffer.getvalue().encode('utf-8')

    def _export_xlsx(self, data: List[Dict], options: ExportOptions) -> bytes:
        """Export data as XLSX."""
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Export"

            if not data:
                buffer = io.BytesIO()
                wb.save(buffer)
                return buffer.getvalue()

            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Write data
            for row_idx, item in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(header))

            # Auto-width columns
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].auto_size = True

            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self._export_csv(data, options)

    def _export_xml(self, data: List[Dict], options: ExportOptions) -> bytes:
        """Export data as XML."""
        import xml.etree.ElementTree as ET

        root = ET.Element("export")
        root.set("version", "2.0.0")
        root.set("created_at", datetime.utcnow().isoformat())

        items = ET.SubElement(root, "items")
        for item_data in data:
            item = ET.SubElement(items, "item")
            for key, value in item_data.items():
                elem = ET.SubElement(item, key)
                elem.text = str(value) if value is not None else ""

        tree = ET.ElementTree(root)
        buffer = io.BytesIO()
        tree.write(buffer, encoding='utf-8', xml_declaration=True)

        return buffer.getvalue()

    def _compress(self, content: bytes, filename: str) -> bytes:
        """Compress content to ZIP."""
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(filename, content)
        buffer.seek(0)
        return buffer.read()

    # ==================== Import Methods ====================

    def import_documents(
        self,
        file_content: bytes,
        format: ImportFormat,
        user_id: uuid.UUID,
        organization_id: Optional[uuid.UUID] = None,
        update_existing: bool = False,
    ) -> ImportResult:
        """
        Import documents from file.

        Args:
            file_content: File content bytes
            format: Import format
            user_id: User performing import
            organization_id: Organization context
            update_existing: Update existing records by ID

        Returns:
            ImportResult with statistics
        """
        # Parse file
        if format == ImportFormat.JSON:
            data = self._parse_json(file_content)
        elif format == ImportFormat.CSV:
            data = self._parse_csv(file_content)
        else:
            raise ValueError(f"Unsupported import format: {format}")

        result = ImportResult(
            success=True,
            total_records=len(data),
            imported=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        for idx, item in enumerate(data):
            try:
                # Validate required fields
                if not item.get("filename"):
                    result.errors.append({
                        "row": idx + 1,
                        "error": "Missing required field: filename",
                    })
                    result.skipped += 1
                    continue

                # Check for existing
                existing = None
                if item.get("id") and update_existing:
                    try:
                        existing = self.db.query(Document).filter(
                            Document.id == uuid.UUID(item["id"]),
                            Document.user_id == user_id,
                        ).first()
                    except ValueError:
                        pass

                if existing:
                    # Update existing
                    existing.filename = item["filename"]
                    if item.get("page_count"):
                        existing.page_count = item["page_count"]
                    result.imported += 1
                else:
                    # Create new (note: actual document would need file upload)
                    result.warnings.append(
                        f"Row {idx + 1}: Document metadata imported, but file upload required"
                    )
                    result.skipped += 1

            except Exception as e:
                result.errors.append({
                    "row": idx + 1,
                    "error": str(e),
                })
                result.skipped += 1

        self.db.commit()
        result.success = len(result.errors) == 0

        return result

    def import_mappings(
        self,
        file_content: bytes,
        format: ImportFormat,
        user_id: uuid.UUID,
    ) -> ImportResult:
        """Import mapping configurations."""
        if format == ImportFormat.JSON:
            data = self._parse_json(file_content)
        elif format == ImportFormat.CSV:
            data = self._parse_csv(file_content)
        else:
            raise ValueError(f"Unsupported import format: {format}")

        result = ImportResult(
            success=True,
            total_records=len(data),
            imported=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        # Import mapping rules
        from backend.models.mapping_profile import MappingProfile

        for idx, item in enumerate(data):
            try:
                if not item.get("name"):
                    result.errors.append({
                        "row": idx + 1,
                        "error": "Missing required field: name",
                    })
                    result.skipped += 1
                    continue

                # Check for existing profile
                existing = self.db.query(MappingProfile).filter(
                    MappingProfile.name == item["name"],
                ).first()

                if existing:
                    # Update
                    if item.get("company_name"):
                        existing.company_name = item["company_name"]
                    if item.get("industry"):
                        existing.industry = item["industry"]
                else:
                    # Create new
                    profile = MappingProfile(
                        name=item["name"],
                        company_name=item.get("company_name"),
                        industry=item.get("industry"),
                    )
                    self.db.add(profile)

                result.imported += 1

            except Exception as e:
                result.errors.append({
                    "row": idx + 1,
                    "error": str(e),
                })
                result.skipped += 1

        self.db.commit()
        result.success = len(result.errors) == 0

        return result

    def restore_backup(
        self,
        file_content: bytes,
        user_id: uuid.UUID,
        merge: bool = False,
    ) -> Dict[str, ImportResult]:
        """
        Restore from backup ZIP file.

        Args:
            file_content: ZIP file content
            user_id: User performing restore
            merge: Merge with existing data instead of replace

        Returns:
            Dictionary of results per entity type
        """
        results = {}

        try:
            buffer = io.BytesIO(file_content)
            with zipfile.ZipFile(buffer, 'r') as zf:
                # Read manifest
                manifest = json.loads(zf.read("manifest.json"))

                # Import each content file
                for filename in manifest.get("contents", []):
                    content = zf.read(filename)

                    if filename == "documents.json":
                        results["documents"] = self.import_documents(
                            content,
                            ImportFormat.JSON,
                            user_id,
                            update_existing=merge,
                        )

        except Exception as e:
            logger.error("backup_restore_failed", error=str(e))
            results["error"] = ImportResult(
                success=False,
                total_records=0,
                imported=0,
                skipped=0,
                errors=[{"error": str(e)}],
                warnings=[],
            )

        return results

    def _parse_json(self, content: bytes) -> List[Dict]:
        """Parse JSON content."""
        data = json.loads(content.decode('utf-8'))
        if isinstance(data, dict):
            # Handle single object or wrapped array
            if "items" in data:
                return data["items"]
            return [data]
        return data

    def _parse_csv(self, content: bytes) -> List[Dict]:
        """Parse CSV content."""
        buffer = io.StringIO(content.decode('utf-8'))
        reader = csv.DictReader(buffer)
        return list(reader)


def get_import_export_service(db: Session) -> ImportExportService:
    """Factory function to get import/export service instance."""
    return ImportExportService(db)
