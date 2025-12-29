"""
Template Populator Service

Populates Excel templates with classified and aggregated financial data.
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.services.gaap_classifier import Classification
from backend.services.template_parser import TemplateStructure

logger = structlog.get_logger(__name__)


class TemplatePopulator:
    """
    Populates Excel templates with financial data.
    
    Key responsibilities:
    - Write aggregated values to correct template cells
    - Inject formulas for calculated rows
    - Handle multi-period columns
    - Preserve template formatting
    """
    
    # Formula definitions for calculated rows
    FORMULAS = {
        "Total Revenue": "=SUM({revenue_items})",
        "Total Cost of Goods Sold": "=SUM({cogs_items})",
        "Gross Profit": "={total_revenue}-{total_cogs}",
        "Total Operating Expenses": "=SUM({opex_items})",
        "Operating Income": "={gross_profit}-{total_opex}",
        "Total Other Income/(Expense)": "=SUM({other_items})",
        "Income Before Income Taxes": "={operating_income}+{other_total}",
        "Net Income": "={income_before_tax}-{tax}",
    }
    
    def __init__(self):
        """Initialize the populator."""
        pass
    
    def _safe_cell_write(self, ws: Worksheet, row: int, col: int, value) -> bool:
        """
        Safely write to a cell, handling merged cells.
        
        Returns True if write succeeded, False if cell was read-only.
        """
        try:
            cell = ws.cell(row=row, column=col)
            cell.value = value
            return True
        except AttributeError as e:
            if "read-only" in str(e):
                # Cell is part of a merged range - find the top-left cell
                for merged_range in ws.merged_cells.ranges:
                    if (row, col) in [(r, c) for r in range(merged_range.min_row, merged_range.max_row + 1) 
                                               for c in range(merged_range.min_col, merged_range.max_col + 1)]:
                        # Write to the top-left cell of the merged range
                        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left.value = value
                        return True
                logger.warning(f"Could not write to cell ({row}, {col}): {e}")
                return False
            raise
    
    def populate(
        self,
        workbook: Workbook,
        structure: TemplateStructure,
        aggregated_data: Dict[str, float],
        classifications: List[Classification],
        periods: List[int],
        company_name: Optional[str] = None,
    ) -> Workbook:
        """
        Populate the template with financial data.
        
        Args:
            workbook: Loaded Excel workbook
            structure: Parsed template structure
            aggregated_data: Dict mapping template rows to values
            classifications: List of classified line items
            periods: List of period years (e.g., [2024])
            company_name: Company name to display
            
        Returns:
            Populated workbook
        """
        ws = workbook.active
        
        logger.info(
            "Populating template",
            aggregated_items=len(aggregated_data),
            periods=periods,
        )
        
        # Set company name
        if company_name:
            self._set_company_name(ws, structure, company_name)
        
        # Set period headers
        if periods:
            self._set_period_headers(ws, structure, periods)
        
        # Clear existing placeholder values in data columns before populating
        data_col = structure.data_start_column
        self._clear_data_columns(ws, structure, data_col)
        
        # Track row ranges for formula injection
        row_ranges = {
            "revenue_items": [],
            "cogs_items": [],
            "opex_items": [],
            "other_items": [],
        }
        
        # Populate aggregated data
        
        for template_row, value in aggregated_data.items():
            row_info = structure.rows.get(template_row)
            
            if row_info:
                row_num = row_info.row_number
                
                # Write value
                cell = ws.cell(row=row_num, column=data_col, value=value)
                
                # Apply number formatting
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                
                # Track for formula ranges
                self._track_row_for_formulas(template_row, row_num, row_ranges, classifications)
                
                logger.debug(f"Populated {template_row} at row {row_num}: {value}")
            else:
                logger.warning(f"Template row not found: {template_row}")
        
        # Inject formulas
        self._inject_formulas(ws, structure, row_ranges, data_col)
        
        return workbook
    
    def _set_company_name(
        self,
        ws: Worksheet,
        structure: TemplateStructure,
        company_name: str
    ) -> None:
        """Set the company name in the template."""
        # Look for "Company Name" placeholder
        for label, row_info in structure.rows.items():
            if "company" in label.lower():
                self._safe_cell_write(ws, row_info.row_number, 2, company_name)
                return
        
        # Default to row 1, column B
        self._safe_cell_write(ws, 1, 2, company_name)
    
    def _set_period_headers(
        self,
        ws: Worksheet,
        structure: TemplateStructure,
        periods: List[int]
    ) -> None:
        """Set period year headers in the data columns."""
        # Find the header row (usually row 4 or 5)
        header_row = 5  # Default
        
        for label, row_info in structure.rows.items():
            if "year" in label.lower() or "ended" in label.lower():
                header_row = row_info.row_number
                break
        
        # Set years in columns starting from C
        for i, year in enumerate(periods):
            col = structure.data_start_column + i
            ws.cell(row=header_row, column=col, value=year)
    
    def _clear_data_columns(
        self,
        ws: Worksheet,
        structure: TemplateStructure,
        data_col: int
    ) -> None:
        """
        Clear all existing values in data columns before populating.
        
        This removes any placeholder values from the template so only
        real aggregated data appears in the output.
        """
        # List of rows that should have their value cells cleared
        # These are the data rows (not headers or section labels)
        data_rows = [
            "Products", "Services",  # Revenue
            "Products", "Services",  # COGS (same names, different section)
            "Research & Development", "Selling, General, and Administrative",  # OpEx
            "Other Income/(Expenses), Net",  # Other
            "Provision for Income Taxes",  # Tax
        ]
        
        # Also clear all numeric cells in the data column
        for label, row_info in structure.rows.items():
            row_num = row_info.row_number
            cell = ws.cell(row=row_num, column=data_col)
            
            # Skip header rows and section labels (only clear data rows)
            skip_labels = ["income statement", "line items", "year", "revenue", 
                          "cost of goods sold", "operating expenses", 
                          "other income/(expenses)", "total", "profit", 
                          "operating income", "income before", "net income"]
            
            is_header = any(skip in label.lower() for skip in skip_labels)
            
            # Clear if it's a data row (has a numeric placeholder)
            if not is_header and cell.value is not None:
                # Only clear numeric values (not formulas we want to keep)
                if isinstance(cell.value, (int, float)):
                    cell.value = None
                    logger.debug(f"Cleared placeholder value at row {row_num}: {label}")
    
    def _track_row_for_formulas(
        self,
        template_row: str,
        row_num: int,
        row_ranges: Dict[str, List[int]],
        classifications: List[Classification]
    ) -> None:
        """Track which rows belong to which formula ranges."""
        template_lower = template_row.lower()
        
        # Check if this is a sub-item (not a total or header)
        is_detail = not any(kw in template_lower for kw in ["total", "profit", "income"])
        
        if not is_detail:
            return
        
        # Determine category from the classifications
        for classification in classifications:
            if classification.template_row == template_row:
                if classification.category == "revenue":
                    row_ranges["revenue_items"].append(row_num)
                elif classification.category == "cost_of_goods_sold":
                    row_ranges["cogs_items"].append(row_num)
                elif classification.category == "operating_expenses":
                    row_ranges["opex_items"].append(row_num)
                elif classification.category == "other_income_expenses":
                    row_ranges["other_items"].append(row_num)
                break
    
    def _inject_formulas(
        self,
        ws: Worksheet,
        structure: TemplateStructure,
        row_ranges: Dict[str, List[int]],
        data_col: int
    ) -> None:
        """Inject Excel formulas for calculated rows."""
        col_letter = get_column_letter(data_col)
        
        # Build formula references
        refs = {}
        
        # Revenue items range
        if row_ranges["revenue_items"]:
            start = min(row_ranges["revenue_items"])
            end = max(row_ranges["revenue_items"])
            refs["revenue_range"] = f"{col_letter}{start}:{col_letter}{end}"
        
        # COGS items range
        if row_ranges["cogs_items"]:
            start = min(row_ranges["cogs_items"])
            end = max(row_ranges["cogs_items"])
            refs["cogs_range"] = f"{col_letter}{start}:{col_letter}{end}"
        
        # OpEx items range
        if row_ranges["opex_items"]:
            start = min(row_ranges["opex_items"])
            end = max(row_ranges["opex_items"])
            refs["opex_range"] = f"{col_letter}{start}:{col_letter}{end}"
        
        # Other items range
        if row_ranges["other_items"]:
            start = min(row_ranges["other_items"])
            end = max(row_ranges["other_items"])
            refs["other_range"] = f"{col_letter}{start}:{col_letter}{end}"
        
        # Get row numbers for formula cells
        rows = structure.rows
        
        # Total Revenue formula
        if "Total Revenue" in rows and refs.get("revenue_range"):
            row_num = rows["Total Revenue"].row_number
            ws.cell(row=row_num, column=data_col, value=f"=SUM({refs['revenue_range']})")
            self._format_total_cell(ws.cell(row=row_num, column=data_col))
            refs["total_revenue"] = f"{col_letter}{row_num}"
        
        # Total COGS formula
        if "Total Cost of Goods Sold" in rows and refs.get("cogs_range"):
            row_num = rows["Total Cost of Goods Sold"].row_number
            ws.cell(row=row_num, column=data_col, value=f"=SUM({refs['cogs_range']})")
            self._format_total_cell(ws.cell(row=row_num, column=data_col))
            refs["total_cogs"] = f"{col_letter}{row_num}"
        
        # Gross Profit formula
        if "Gross Profit" in rows and refs.get("total_revenue") and refs.get("total_cogs"):
            row_num = rows["Gross Profit"].row_number
            ws.cell(row=row_num, column=data_col, value=f"={refs['total_revenue']}-{refs['total_cogs']}")
            self._format_calculated_cell(ws.cell(row=row_num, column=data_col))
            refs["gross_profit"] = f"{col_letter}{row_num}"
        
        # Total Operating Expenses formula
        if "Total Operating Expenses" in rows and refs.get("opex_range"):
            row_num = rows["Total Operating Expenses"].row_number
            ws.cell(row=row_num, column=data_col, value=f"=SUM({refs['opex_range']})")
            self._format_total_cell(ws.cell(row=row_num, column=data_col))
            refs["total_opex"] = f"{col_letter}{row_num}"
        
        # Operating Income formula
        if "Operating Income" in rows and refs.get("gross_profit") and refs.get("total_opex"):
            row_num = rows["Operating Income"].row_number
            ws.cell(row=row_num, column=data_col, value=f"={refs['gross_profit']}-{refs['total_opex']}")
            self._format_calculated_cell(ws.cell(row=row_num, column=data_col))
            refs["operating_income"] = f"{col_letter}{row_num}"
        
        # Total Other Income/(Expense) formula
        if "Total Other Income/(Expense)" in rows and refs.get("other_range"):
            row_num = rows["Total Other Income/(Expense)"].row_number
            ws.cell(row=row_num, column=data_col, value=f"=SUM({refs['other_range']})")
            self._format_total_cell(ws.cell(row=row_num, column=data_col))
            refs["other_total"] = f"{col_letter}{row_num}"
        elif "Total Other Income/(Expense)" in rows:
            # Even without other_range, set the reference for downstream formulas
            refs["other_total"] = f"{col_letter}{rows['Total Other Income/(Expense)'].row_number}"
        
        # Income Before Income Taxes formula
        # Build formula from available components (operating income + other total)
        if "Income Before Income Taxes" in rows:
            row_num = rows["Income Before Income Taxes"].row_number
            
            # Get operating income reference (prefer existing ref, fall back to row)
            if refs.get("operating_income"):
                operating = refs["operating_income"]
            elif "Operating Income" in rows:
                operating = f"{col_letter}{rows['Operating Income'].row_number}"
                refs["operating_income"] = operating
            else:
                operating = "0"
            
            # Get other total reference (prefer existing ref, fall back to row)
            if refs.get("other_total"):
                other = refs["other_total"]
            elif "Total Other Income/(Expense)" in rows:
                other = f"{col_letter}{rows['Total Other Income/(Expense)'].row_number}"
                refs["other_total"] = other
            else:
                other = "0"
            
            ws.cell(row=row_num, column=data_col, value=f"={operating}+{other}")
            self._format_calculated_cell(ws.cell(row=row_num, column=data_col))
            refs["income_before_tax"] = f"{col_letter}{row_num}"
        
        # Net Income formula
        if "Net Income" in rows:
            row_num = rows["Net Income"].row_number
            
            # Get income before tax reference
            if refs.get("income_before_tax"):
                income_before_tax = refs["income_before_tax"]
            elif "Income Before Income Taxes" in rows:
                income_before_tax = f"{col_letter}{rows['Income Before Income Taxes'].row_number}"
            else:
                income_before_tax = "0"
            
            tax_row = rows.get("Provision for Income Taxes")
            tax = f"{col_letter}{tax_row.row_number}" if tax_row else "0"
            ws.cell(row=row_num, column=data_col, value=f"={income_before_tax}-{tax}")
            self._format_net_income_cell(ws.cell(row=row_num, column=data_col))
    
    def _format_total_cell(self, cell) -> None:
        """Apply formatting for total rows."""
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
    
    def _format_calculated_cell(self, cell) -> None:
        """Apply formatting for calculated rows (Gross Profit, etc.)."""
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
    
    def _format_net_income_cell(self, cell) -> None:
        """Apply formatting for Net Income row (double underline)."""
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='double'),
        )


# Singleton instance
_populator_instance: Optional[TemplatePopulator] = None


def get_template_populator() -> TemplatePopulator:
    """Get singleton TemplatePopulator instance."""
    global _populator_instance
    if _populator_instance is None:
        _populator_instance = TemplatePopulator()
    return _populator_instance
