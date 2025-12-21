"""
Excel parser service.

Parses Excel files to extract cells, formulas, named ranges, and metadata.
"""
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import structlog

logger = structlog.get_logger(__name__)

# Try to import openpyxl
try:
    from openpyxl import load_workbook
    from openpyxl.cell import Cell
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed, Excel parser disabled")


@dataclass
class ParsedCell:
    """Represents a parsed cell from Excel."""

    sheet: str
    row: int
    column: int
    address: str  # e.g., "A1"
    value: Any
    formula: Optional[str] = None
    data_type: str = "n"  # n=numeric, s=string, b=boolean, etc.
    is_merged: bool = False
    number_format: Optional[str] = None
    font_bold: bool = False
    indent: int = 0


@dataclass
class ParsedSheet:
    """Represents a parsed worksheet."""

    name: str
    max_row: int
    max_column: int
    cells: List[ParsedCell] = field(default_factory=list)
    merged_ranges: List[str] = field(default_factory=list)


@dataclass
class NamedRange:
    """Represents a named range in Excel."""

    name: str
    scope: Optional[str]  # Sheet name or None for global
    addresses: List[str]


@dataclass
class ParsedWorkbook:
    """Represents a fully parsed Excel workbook."""

    filename: str
    sheets: List[ParsedSheet]
    named_ranges: List[NamedRange]
    active_sheet: Optional[str]
    metadata: Dict[str, Any] = field(default_factory=dict)


class ExcelParser:
    """
    Service for parsing Excel files.

    Extracts all cell data, formulas, named ranges, and structural information.
    """

    # Common formula functions to detect calculated cells
    FORMULA_FUNCTIONS = {
        "SUM", "AVERAGE", "COUNT", "IF", "VLOOKUP", "HLOOKUP",
        "INDEX", "MATCH", "SUMIF", "COUNTIF", "MAX", "MIN",
        "ROUND", "ABS", "IFERROR", "EOMONTH", "YEAR", "MONTH",
    }

    def __init__(self):
        """Initialize Excel parser."""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available")

    def parse(self, file_path: Path) -> ParsedWorkbook:
        """
        Parse an Excel file.

        Args:
            file_path: Path to Excel file.

        Returns:
            ParsedWorkbook with all extracted data.
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed")

        logger.info("Parsing Excel file", path=str(file_path))

        wb = load_workbook(filename=str(file_path), data_only=False)

        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parsed_sheet = self._parse_sheet(ws)
            sheets.append(parsed_sheet)

        named_ranges = self._parse_named_ranges(wb)

        result = ParsedWorkbook(
            filename=file_path.name,
            sheets=sheets,
            named_ranges=named_ranges,
            active_sheet=wb.active.title if wb.active else None,
            metadata={
                "sheet_count": len(sheets),
                "created": str(wb.properties.created) if wb.properties.created else None,
                "modified": str(wb.properties.modified) if wb.properties.modified else None,
            },
        )

        logger.info(
            "Excel file parsed",
            sheets=len(sheets),
            named_ranges=len(named_ranges),
        )

        wb.close()
        return result

    def _parse_sheet(self, ws: "Worksheet") -> ParsedSheet:
        """
        Parse a single worksheet.

        Args:
            ws: openpyxl Worksheet object.

        Returns:
            ParsedSheet with all cells.
        """
        cells = []
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        merged_cells_set = set()

        # Build set of merged cell addresses
        for mr in ws.merged_cells.ranges:
            for cell in mr.cells:
                row, col = cell
                merged_cells_set.add((row, col))

        # Parse all cells
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Skip None cells in sparse worksheets
                if cell.value is None and cell.data_type == "n" and not cell.font.bold:
                    continue

                address = f"{get_column_letter(col_idx)}{row_idx}"

                # Get formula if present
                formula = None
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    formula = str(cell.value)
                elif hasattr(cell, 'value') and cell.value is not None:
                    # Check if the cell has a cached formula
                    pass

                parsed_cell = ParsedCell(
                    sheet=ws.title,
                    row=row_idx,
                    column=col_idx,
                    address=address,
                    value=cell.value if not (isinstance(cell.value, str) and cell.value.startswith("=")) else None,
                    formula=formula,
                    data_type=cell.data_type,
                    is_merged=(row_idx, col_idx) in merged_cells_set,
                    number_format=cell.number_format,
                    font_bold=cell.font.bold if cell.font else False,
                    indent=cell.alignment.indent if cell.alignment and cell.alignment.indent else 0,
                )
                cells.append(parsed_cell)

        return ParsedSheet(
            name=ws.title,
            max_row=ws.max_row,
            max_column=ws.max_column,
            cells=cells,
            merged_ranges=merged_ranges,
        )

    def _parse_named_ranges(self, wb) -> List[NamedRange]:
        """
        Parse named ranges from workbook.

        Args:
            wb: openpyxl Workbook object.

        Returns:
            List of NamedRange objects.
        """
        named_ranges = []

        for name in wb.defined_names.definedName:
            try:
                destinations = list(name.destinations)
                addresses = []
                scope = None

                for sheet, coord in destinations:
                    addresses.append(f"{sheet}!{coord}")
                    scope = sheet

                named_ranges.append(NamedRange(
                    name=name.name,
                    scope=scope if len(destinations) == 1 else None,
                    addresses=addresses,
                ))
            except Exception as e:
                logger.debug("Failed to parse named range", name=name.name, error=str(e))

        return named_ranges

    def get_cell_references(self, formula: str) -> List[str]:
        """
        Extract cell references from a formula.

        Args:
            formula: Excel formula string.

        Returns:
            List of cell addresses referenced.
        """
        if not formula or not formula.startswith("="):
            return []

        # Pattern for cell references: A1, $A$1, Sheet1!A1, etc.
        pattern = r"(?:\'?[\w\s]+\'?!)?\$?[A-Za-z]+\$?\d+"
        matches = re.findall(pattern, formula)

        return matches

    def is_formula_cell(self, cell: ParsedCell) -> bool:
        """Check if a cell contains a formula."""
        return cell.formula is not None and cell.formula.startswith("=")

    def is_input_cell(self, cell: ParsedCell) -> bool:
        """Check if a cell is likely an input cell (user enters data)."""
        if cell.formula:
            return False
        if cell.value is None:
            return False
        if isinstance(cell.value, str) and not cell.value.strip():
            return False
        return True


# Singleton instance
_parser_instance: Optional[ExcelParser] = None


def get_excel_parser() -> ExcelParser:
    """Get singleton ExcelParser instance."""
    global _parser_instance
    if _parser_instance is None:
        _parser_instance = ExcelParser()
    return _parser_instance
